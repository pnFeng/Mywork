import time
from ICs import PineHurst
from openpyxl import Workbook
from bit_manipulate import bm
from Mylib.bench_resource import brsc


class PHCurMeas:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

    def enable(self):
        # enable
        ret, data = self.dsh.i2c_read_reg(0x2F)
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)
        self.dsh.i2c_write_reg(0x32, 0x80)

    def CurMeas(self, power, acc, rail):
        sum = 0
        if rail == 'a':
            rgt = 0x0C
        elif rail == 'b':
            rgt = 0x0E
        elif rail == 'c':
            rgt = 0x0F
        if acc == 0:
            ret, data = self.dsh.i2c_read_reg(0x6C)
            data = bm.clear_bit(data, 4)
            self.dsh.i2c_write_reg(0x6C, data)
            if power == 0:
                ret, data = self.dsh.i2c_read_reg(0x1B)
                data = bm.clear_bit(data, 6)
                self.dsh.i2c_write_reg(0x1B, data)
            elif power == 1:
                ret, data = self.dsh.i2c_read_reg(0x1B)
                data = bm.set_bit(data, 6)
                self.dsh.i2c_write_reg(0x1B, data)
        elif acc == 1:
            ret, data = self.dsh.i2c_read_reg(0x6C)
            data = bm.set_bit(data, 4)
            self.dsh.i2c_write_reg(0x6C, data)
            if power == 0:
                ret, data = self.dsh.i2c_read_reg(0x1B)
                data = bm.clear_bit(data, 6)
                self.dsh.i2c_write_reg(0x1B, data)
            elif power == 1:
                ret, data = self.dsh.i2c_read_reg(0x1B)
                data = bm.set_bit(data, 6)
                self.dsh.i2c_write_reg(0x1B, data)
        for i in range(10):
            ret, data = self.dsh.i2c_read_reg(rgt)
            ret, data1 = self.dsh.i2c_read_reg(0x2E)
            if acc == 1:
                if rail == 'a':
                    data = (data << 1) + bm.get_bit(data1, 'l', [7])[0]
                elif rail == 'b':
                    data = (data << 1) + bm.get_bit(data1, 'l', [5])[0]
                elif rail == 'c':
                    data = (data << 1) + bm.get_bit(data1, 'l', [4])[0]

            sum = data + sum
            time.sleep(0.1)
            avag = round(sum / 10)
        return avag


Rail = 'b'
path = '../CurrentADC_' + Rail + '.xlsx'
step = 0.125
precise_step = 0.015625
wb = Workbook()
ws1 = wb.active
ws1.cell(row=1, column=1, value='load')
ws1.cell(row=1, column=2, value='ADC_6bit')
ws1.cell(row=1, column=3, value='ADC_6bit_Current')
ws1.cell(row=1, column=4, value='Delta_LSB')
ws1.cell(row=1, column=5, value='ADC_9bit')
ws1.cell(row=1, column=6, value='ADC_9bit_Current')
ws1.cell(row=1, column=7, value='Delta_9bit_LSB')
ws1.cell(row=1, column=8, value='Real Power')
ws1.cell(row=1, column=9, value='ADC_value_Power')
ws1.cell(row=1, column=10, value='ADC_Power')
ws1.cell(row=1, column=11, value='Delta_power_LSB')
ws1.cell(row=1, column=12, value='ADCpower_precise')
ws1.cell(row=1, column=13, value='Delta_power_precise_LSB')

el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
p1.on(5, 2, 1)
p = PHCurMeas()
p.enable()
time.sleep(1)

for i in range(401):
    if i == 0:
        load = 0
    else:
        load += 0.01
    el1.cc(load, 3)
    el1.on([3])
    time.sleep(0.5)
    load1 = el1.curmeas([1])
    Actual_Power = 1.1 * load1

    Reading0 = p.CurMeas(0, 0, Rail)
    Current = Reading0 * step
    Current_LSB = Reading0 - round(load1 / step)

    Reading1 = p.CurMeas(1, 0, Rail)
    power = Reading1 * step
    power_LSB = Reading1 - round(Actual_Power / step)

    Reading2 = p.CurMeas(0, 1, Rail)
    Current_precise = Reading2 * precise_step
    Current_precise_LSB = Reading2 - round(load1 / precise_step)

    Reading3 = p.CurMeas(1, 1, Rail)
    power_precise = Reading3 * precise_step
    power_precise_LSB = Reading3 - round(Actual_Power / precise_step)

    ws1.cell(row=i + 2, column=1, value=load1)
    ws1.cell(row=i + 2, column=2, value=hex(int(Reading0)))
    ws1.cell(row=i + 2, column=3, value=Current)
    ws1.cell(row=i + 2, column=4, value=Current_LSB)
    ws1.cell(row=i + 2, column=5, value=hex(int(Reading2)))
    ws1.cell(row=i + 2, column=6, value=Current_precise)
    ws1.cell(row=i + 2, column=7, value=Current_precise_LSB)
    ws1.cell(row=i + 2, column=8, value=Actual_Power)
    ws1.cell(row=i + 2, column=9, value=hex(int(Reading1)))
    ws1.cell(row=i + 2, column=10, value=power)
    ws1.cell(row=i + 2, column=11, value=power_LSB)
    ws1.cell(row=i + 2, column=12, value=power_precise)
    ws1.cell(row=i + 2, column=13, value=power_precise_LSB)

    print(i, Reading2, Current_precise_LSB, load1)

time.sleep(1)
el1.off([1])
wb.save(path)
