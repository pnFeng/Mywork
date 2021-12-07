import time
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from Mylib.bench_resource import brsc
from ICs import PineHurst
from bit_manipulate import bm


class Piht:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

    def enable(self, bEnable):
        if bEnable:
            self.dsh.i2c_write_reg(0x32, 0x80)
        else:
            self.dsh.i2c_write_reg(0x32, 0x00)


path = './../' + 'LoadRegulation.xlsx'
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
s1 = brsc.Scope('USB0::0x05FF::0x1023::3808N60392::INSTR')
m1 = brsc.DMM('USB0::0x2A8D::0x1401::MY57215694::0::INSTR')
m2 = brsc.DMM('USB0::0x2A8D::0x1401::MY57218437::0::INSTR')
p1.on(5, 2, 1)

wb = Workbook()
ws1 = wb.active
loadchannel = 2

# a = Piht()
# a.dsh.i2c_write_reg(0x50, 0xcf)
# a.dsh.i2c_write_reg(0x56, 0x0f)
# a.dsh.i2c_write_reg(0x2F, 0x42)
# a.enable(True)

group = 0
i, j = 0, 0
max = 5
while True:
    j += 1
    if j == 1:
        i = 0
    else:
        i += 5
    if i >= 101:
        break
    load = round((i * max) / 100, 3)
    print(i, load)

    el1.cc(load, loadchannel)
    el1.on([loadchannel])
    time.sleep(0.5)

    iout = el1.curmeas(loadchannel)
    v1 = m1.meas()

    ws1.cell(row=j + 1, column=1 + 8 * group, value=iout)
    ws1.cell(row=j + 1, column=2 + 8 * group, value=v1)

    ws1.cell(row=1, column=1 + 8 * group, value='IOUT')
    ws1.cell(row=1, column=2 + 8 * group, value='v1')

wb.save(path)

el1.off([loadchannel])
