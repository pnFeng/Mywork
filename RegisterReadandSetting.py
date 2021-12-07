import sys
import time
from openpyxl import Workbook, load_workbook
# import dolphin
from ICs import DolphinSH
from bit_manipulate import bm
from Mylib.bench_resource import brsc


class Piht:
    def __init__(self):
        self.dsh = DolphinSH()
        ret = self.dsh.connect()
        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=0)
        if ret is False:
            print("Failed to connect")
            sys.exit(-1)

        # self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

    def unlockDIMM(self, bEnable):
        if bEnable == 1:
            # # Un-lock DIMM vendor region(0x40~0x6f)
            self.dsh.i2c_write_reg(0x37, 0x73)
            self.dsh.i2c_write_reg(0x38, 0x94)
            self.dsh.i2c_write_reg(0x39, 0x40)

    def unlockIDT(self, bEnable):
        if bEnable == 1:
            # # Un-lock PMIC vendor region(0x70~0xff)
            self.dsh.i2c_write_reg(0x37, 0x79)
            self.dsh.i2c_write_reg(0x38, 0xbe)
            self.dsh.i2c_write_reg(0x39, 0x10)

    def securemode(self, bEnable):
        ret, data = self.dsh.i2c_read_reg(0x2F)
        if bEnable == 0:
            data = bm.set_bit(data, 2)
        elif bEnable == 1:
            data = bm.clear_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

    def enable(self, bEnable):
        if bEnable:
            self.dsh.i2c_write_reg(0x32, 0x80)
        else:
            self.dsh.i2c_write_reg(0x32, 0x00)


wb = load_workbook("../SKReg.xlsx")
sheet = wb.active
reg, data, errval = [], [], []
for i in range(104):
    reg_temp = sheet.cell(i + 2, 1).value
    val_temp = sheet.cell(i + 2, 2).value
    reg.append(reg_temp)
    data.append(val_temp)

a = Piht()

a.securemode(0)
a.unlockDIMM(1)

# for i in range(len(reg)):
#     a.dsh.i2c_write_reg(int(reg[i], 16), int(data[i], 16))

for i in range(len(reg)):
    ret, val = a.dsh.i2c_read_reg(int(reg[i], 16))
    if val != int(data[i], 16):
        errval.append((reg[i], hex(val)))

if len(errval) != 0:
    print('discordance {}'.format(errval))
else:
    print('OK')

# a.dsh.i2c_write_reg(0x39, 0x74)  # Burning 0x00~0x3F
# time.sleep(1)
# a.dsh.i2c_write_reg(0x39, 0x81)  # Burning 0x40~0x4F
# time.sleep(1)
# a.dsh.i2c_write_reg(0x39, 0x82)  # Burning 0x50~0x5F
# time.sleep(1)
# a.dsh.i2c_write_reg(0x39, 0x85)  # Burning 0x60~0x6F
# time.sleep(1)
