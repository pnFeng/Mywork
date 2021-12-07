import time
from ICs import PineHurst
from openpyxl import Workbook
from bit_manipulate import bm
from Mylib.bench_resource import brsc


class PHCurMeas:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

        # 0x1A[1]=1
        ret, data = self.dsh.i2c_read_reg(0x1A)
        assert ret == 0
        data = bm.set_bit(data, 1)
        self.dsh.i2c_write_reg(0x1A, data)

        # 0x1b[6]=1
        ret, data = self.dsh.i2c_read_reg(0x1B)
        assert ret == 0
        data = bm.set_bit(data, 6)
        self.dsh.i2c_write_reg(0x1B, data)

    def enable(self):
        # enable
        ret, data = self.dsh.i2c_read_reg(0x2F)
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)
        self.dsh.i2c_write_reg(0x32, 0x80)

    def PowerMeas(self):
        sum = 0

        for i in range(10):
            ret, data = self.dsh.i2c_read_reg(0x0C)
            sum = data + sum
            time.sleep(0.1)
        avag = round(sum / 10)
        return avag


path = '../TotalPowerADC.xlsx'
step = 0.125
loadA = 4
loadB = 4
loadC = 1
wb = Workbook()
ws1 = wb.active
ws1.cell(row=1, column=1, value='Item')
ws1.cell(row=1, column=2, value='Total Power')
ws1.cell(row=1, column=3, value='Total Power ADC')
ws1.cell(row=1, column=4, value='LSB')

el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
p1.on(5, 4, 1)
p = PHCurMeas()
p.enable()
time.sleep(1)

n = 100
for i in range(n + 1):
    if i == 0:
        A, B, C = 0, 0, 0
    else:
        A = loadA * i * 1 / n
        B = loadB * i * 1 / n
        C = loadC * i * 1 / n
    print(A, B, C)

    el1.cc(A, 1)
    el1.cc(C, 2)
    el1.cc(B, 3)
    time.sleep(0.5)

    el1.on([1, 2, 3])

    time.sleep(0.5)
    powerA = 1.1 * el1.curmeas(1)
    powerB = 1.1 * el1.curmeas(3)
    powerC = 1.8 * el1.curmeas(2)

    TotalPower = powerA + powerB + powerC
    ADC = p.PowerMeas()
    ADCPower = ADC * step
    LSB = ADC - round(TotalPower / step)

    ws1.cell(row=i + 2, column=1, value=i)
    ws1.cell(row=i + 2, column=2, value=TotalPower)
    ws1.cell(row=i + 2, column=3, value=ADCPower)
    ws1.cell(row=i + 2, column=4, value=LSB)

    print(i, powerA, powerB, powerC, ADCPower)

time.sleep(1)
el1.off([1])
wb.save(path)
