import time
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from Mylib.bench_resource import brsc
from ICs import PineHurst
from bit_manipulate import bm


class Piht:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

    def enable(self, bEnable):
        if bEnable:
            self.dsh.i2c_write_reg(0x32, 0x80)
        else:
            self.dsh.i2c_write_reg(0x32, 0x00)


path = './../' + 'LineRegulation.xlsx'
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
s1 = brsc.Scope('USB0::0x05FF::0x1023::3808N60392::INSTR')
m1 = brsc.DMM('USB0::0x2A8D::0x1401::MY57215694::0::INSTR')
m2 = brsc.DMM('USB0::0x2A8D::0x1401::MY57218437::0::INSTR')
p1.on(5, 2, 1)

wb = Workbook()
ws1 = wb.active
loadchannel = 4

# a = Piht()
# a.dsh.i2c_write_reg(0x50, 0xcf)
# a.dsh.i2c_write_reg(0x56, 0x0f)
# a.dsh.i2c_write_reg(0x2F, 0x42)
# a.enable(True)
vin = [4.25, 4.5, 4.75, 5, 5.25, 5.5]
full = 0.05

for i in vin:
    p1.on(i, 2, 1)
    for j in range(4):
        if j == 0:
            load = 0
        elif j == 1:
            load = 0.1 * full
        elif j == 2:
            load = 0.5 * full
        elif j == 3:
            load = full
        el1.cc(load, loadchannel)
        el1.on([loadchannel])
        print(i, load)
        time.sleep(0.5)
        v1 = m1.meas()

        ws1.cell(row=vin.index(i) + 2, column=1, value=i)
        ws1.cell(row=vin.index(i) + 2, column=j + 2, value=v1)

    ws1.cell(row=1, column=1, value='VIN')
    ws1.cell(row=1, column=2, value='VOUT_0')
    ws1.cell(row=1, column=3, value='VOUT_10%')
    ws1.cell(row=1, column=4, value='VOUT_50%')
    ws1.cell(row=1, column=5, value='VOUT_100%')

wb.save(path)

el1.off([loadchannel])
