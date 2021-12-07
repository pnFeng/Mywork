import time
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from Mylib.bench_resource import brsc
from ICs import PineHurst
from bit_manipulate import bm
import random
from openpyxl.styles import colors, Font


class Piht:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

    def enable(self, bEnable):
        if bEnable:
            self.dsh.i2c_write_reg(0x32, 0x80)
        else:
            self.dsh.i2c_write_reg(0x32, 0x00)


path = './../' + 'EFF.xlsx'
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
m1 = brsc.DMM('USB0::0x2A8D::0x1401::MY57215694::0::INSTR')
m2 = brsc.DMM('USB0::0x2A8D::0x1401::MY57218437::0::INSTR')
p1.on(5, 2, 1)

wb = Workbook()
ws1 = wb.active

a = Piht()
# a.dsh.i2c_write_reg(0x4f, 0x01)
# a.dsh.i2c_write_reg(0xac, 0x63)
# a.dsh.i2c_write_reg(0x8a, 0x1e)
# a.dsh.i2c_write_reg(0x92, 0x77)
# a.dsh.i2c_write_reg(0x2F, 0x42)
# a.enable(True)

# a.dsh.i2c_write_reg(0x2F, 0x42)

## Program/Security Mode
# a.dsh.i2c_write_reg(0x2f, 0x06)
#
# # Maxmize OCP
# a.dsh.i2c_write_reg(0x20, 0xCF)
#
# # loop compensation
# a.dsh.i2c_write_reg(0x80, 0x09)
# a.dsh.i2c_write_reg(0x88, 0x01)
# a.dsh.i2c_write_reg(0xac, 0xa2)
#
# # deadtime
# a.dsh.i2c_write_reg(0x89, 0xa4)
#
# # drive strength
# a.dsh.i2c_write_reg(0x86, 0xF8)
# a.dsh.i2c_write_reg(0x87, 0xD7)
#
# # Adjust ILIM & ZCD
# # write 0x90
# a.dsh.i2c_write_reg(0x91, 0x89)
#
# # write 0x98
# # write 0x99 0x15
#
# # write 0xa6
# a.dsh.i2c_write_reg(0xa7, 0x09)
#
# # FSW
# a.dsh.i2c_write_reg(0x94, 0x07)
# a.dsh.i2c_write_reg(0x29, 0x90)  # 750kHz 1MHz
# a.dsh.i2c_write_reg(0x2A, 0x88)  # 750kHz 1MHz
# # write 0xa2 0x06
# a.dsh.i2c_write_reg(0xaa, 0x06)
#
# # Output Voltage
# a.dsh.i2c_write_reg(0x93, 0x03)
# a.dsh.i2c_write_reg(0xa9, 0x05)

# 500kHz
# a.dsh.i2c_write_reg(0x84, 0x25)
# a.dsh.i2c_write_reg(0x94, 0x05)
# Enable
# a.dsh.i2c_write_reg(0xb2, 0x00)
# a.dsh.i2c_write_reg(0x32, 0x80)

Rail = 'C'
if Rail == 'A':
    a.dsh.i2c_write_reg(0x2f, 0x42)  # RAILA
    channel = 1
    max = 4
elif Rail == 'B':
    a.dsh.i2c_write_reg(0x2f, 0x12)  # RAILB
    channel = 1
    max = 4
elif Rail == 'C':
    a.dsh.i2c_write_reg(0x2f, 0x0A)  # RAILC
    channel = 1
    max = 1

step1 = 0.1
step2 = 0.25
group = 0
maxcell = 0
i, j = -step1, 0

while True:
    j += 1
    # i += 5
    # if i >= 101:
    #     break
    # load = round((i * max) / 100, 2)
    i = round(i, 2)
    if i < 1:
        i += step1
    else:
        i += step2
    if i > max:
        break
    load = i
    print(i, load)

    el1.cc(load, channel)
    el1.on([channel])
    time.sleep(0.5)

    iout = el1.curmeas(channel)
    v1 = m1.meas()
    v2 = m2.meas()
    Iin_p = p1.curmeas(1)

    v1 = p1.volmeas(1)
    # v1 = v1 - Iin_p * (0.02 + 0.01 * random.randint(1, 2))

    power_in = v1 * Iin_p
    power_out = v2 * iout
    Eff = 100 * power_out / power_in
    if maxcell < Eff:
        maxcell = Eff
        red_c = 5 + 8 * group
        red_r = j + 1

    ws1.cell(row=j + 1, column=1 + 8 * group, value=v1)
    ws1.cell(row=j + 1, column=2 + 8 * group, value=Iin_p)
    ws1.cell(row=j + 1, column=3 + 8 * group, value=v2)
    ws1.cell(row=j + 1, column=4 + 8 * group, value=iout)
    ws1.cell(row=j + 1, column=5 + 8 * group, value=Eff)

    ws1.cell(row=1, column=1 + 8 * group, value='Vin_Bulk')
    ws1.cell(row=1, column=2 + 8 * group, value='Iin_Bulk')
    ws1.cell(row=1, column=3 + 8 * group, value='VOUT')
    ws1.cell(row=1, column=4 + 8 * group, value='IOUT')
    ws1.cell(row=1, column=5 + 8 * group, value='Efficiency')

ft = Font(color=colors.RED, size=12)
c = ws1.cell(row=red_r, column=red_c)
c.font = ft

wb.save(path)

el1.off([channel])
