import time
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from Mylib.bench_resource import brsc
from ICs import PineHurst
from bit_manipulate import bm


class Piht:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        # Program Mode
        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

    def enable(self, bEnable):
        if bEnable:
            self.dsh.i2c_write_reg(0x32, 0x80)
        else:
            self.dsh.i2c_write_reg(0x32, 0x00)

    def sleepmode(self, bEnable):
        if bEnable:
            ret, data = self.dsh.i2c_read_reg(0x88)
            data = bm.set_bit(data, 6)
            self.dsh.i2c_write_reg(0x88, data)
        else:
            ret, data = self.dsh.i2c_read_reg(0x88)
            data = bm.clear_bit(data, 6)
            self.dsh.i2c_write_reg(0x88, data)

    def Curbal(self, bEnable):
        if bEnable:
            ret, data = self.dsh.i2c_read_reg(0x88)
            data = bm.set_bit(data, 7)
            self.dsh.i2c_write_reg(0x88, data)
        else:
            ret, data = self.dsh.i2c_read_reg(0x88)
            data = bm.clear_bit(data, 7)
            self.dsh.i2c_write_reg(0x88, data)


path = './../' + 'EFF_Dualphase.xlsx'
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
m1 = brsc.DMM('USB0::0x2A8D::0x1401::MY57215694::0::INSTR')
m2 = brsc.DMM('USB0::0x2A8D::0x1401::MY57218437::0::INSTR')
p1.on(5, 3, 1)

wb = Workbook()
ws1 = wb.active

a = Piht()
a.dsh.i2c_write_reg(0x4f, 0x01)
# a.dsh.i2c_write_reg(0xac, 0x63)
# a.dsh.i2c_write_reg(0x8a, 0x02)
# a.dsh.i2c_write_reg(0x92, 0x87)
# a.dsh.i2c_write_reg(0x56, 0xcf)
# a.dsh.i2c_write_reg(0x2F, 0x42)
# a.enable(True)

a.sleepmode(1)
a.Curbal(0)
ret, data = a.dsh.i2c_read_reg(0x88)

a.dsh.i2c_write_reg(0x2F, 0x42)
# a.dsh.i2c_write_reg(0x2F, 0x0A)

step1 = 0.1
step2 = 0.25
group = 0
i, j = -step1, 0
max = 8

while True:
    j += 1
    # i += 5
    # if i >= 101:
    #     break
    # load = round((i * max) / 100, 2)
    i = round(i, 2)
    if i < 1:
        i += step1
    else:
        i += step2
    if i > max:
        break
    load = i
    print(i, load)

    el1.cc(load, 1)
    el1.on([1])
    time.sleep(0.5)

    iout = el1.curmeas(1)
    v1 = m1.meas()
    v2 = m2.meas()
    Iin_p = p1.curmeas(1)

    power_in = v1 * Iin_p
    power_out = v2 * iout
    Eff = 100 * power_out / power_in

    ws1.cell(row=j + 1, column=1 + 8 * group, value=v1)
    ws1.cell(row=j + 1, column=2 + 8 * group, value=Iin_p)
    ws1.cell(row=j + 1, column=3 + 8 * group, value=v2)
    ws1.cell(row=j + 1, column=4 + 8 * group, value=iout)
    ws1.cell(row=j + 1, column=5 + 8 * group, value=Eff)

    ws1.cell(row=1, column=1 + 8 * group, value='Vin_Bulk')
    ws1.cell(row=1, column=2 + 8 * group, value='Iin_Bulk')
    ws1.cell(row=1, column=3 + 8 * group, value='VOUT')
    ws1.cell(row=1, column=4 + 8 * group, value='IOUT')
    ws1.cell(row=1, column=5 + 8 * group, value='Efficiency')

wb.save(path)

el1.off([1])
