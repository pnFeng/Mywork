import time
from openpyxl import Workbook
from ICs import PineHurst
from Mylib.bench_resource import brsc
from bit_manipulate import bm


class PiHs:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

    def enable(self, bEnable):
        if bEnable:
            self.dsh.i2c_write_reg(0x32, 0x80)
        else:
            self.dsh.i2c_write_reg(0x32, 0x00)

    def ADCConfig(self, ch):
        if ch == 'a':
            self.dsh.i2c_write_reg(0x30, 0x80)
        elif ch == 'b':
            self.dsh.i2c_write_reg(0x30, 0x90)
        elif ch == 'c':
            self.dsh.i2c_write_reg(0x30, 0x98)

    def ADCrslt(self):
        ret, data1 = self.dsh.i2c_read_reg(0x31)
        ret, data2 = self.dsh.i2c_read_reg(0x72)
        data2 = data2 & 0x03
        return data1, data2

    def enableFLT(self, enable):
        if enable == 1:
            ret, data = self.dsh.i2c_read_reg(0xc7)
            data = data | 0x40
            self.dsh.i2c_write_reg(0xc7, data)
            ret, data = self.dsh.i2c_read_reg(0xc7)
            print('0xc7_1', hex(data))
        elif enable == 0:
            ret, data = self.dsh.i2c_read_reg(0xc7)
            data = data & 0xBF
            self.dsh.i2c_write_reg(0xc7, data)
            ret, data = self.dsh.i2c_read_reg(0xc7)
            print('0xc7_0', hex(data))


path = './../' + 'ADC.xlsx'
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
m1 = brsc.DMM('USB0::0x2A8D::0x1401::MY57215694::0::INSTR')
m2 = brsc.DMM('USB0::0x2A8D::0x1401::MY57218437::0::INSTR')
p1.on(5, 4, 1)

wb = Workbook()
ws1 = wb.active

a = PiHs()
a.ADCConfig('c')
a.enable(1)

el1.cc(4, 1)
el1.cc(1, 2)
el1.cc(4, 3)
el1.on([1, 2, 3])

ws1.cell(1, 1, '#')
ws1.cell(1, 2, 'Time')
ws1.cell(1, 3, 'Multimeter VOUT(V)')
ws1.cell(1, 4, 'Target ADC_8bit')
ws1.cell(1, 5, 'Target ADC_10bit')
ws1.cell(1, 6, 'ADC_8bit')
ws1.cell(1, 7, 'ADC_10bit')
ws1.cell(1, 8, 'Delta_8bit')
ws1.cell(1, 9, 'Delta_10bit')

ws1.cell(1, 13, 'Time')
ws1.cell(1, 14, 'Multimeter VOUT(V)')
ws1.cell(1, 15, 'Target ADC_8bit')
ws1.cell(1, 16, 'Target ADC_10bit')
ws1.cell(1, 17, 'ADC_8bit')
ws1.cell(1, 18, 'ADC_10bit')
ws1.cell(1, 19, 'Delta_8bit')
ws1.cell(1, 20, 'Delta_10bit')

n = 10
a.enableFLT(0)
for i in range(n):
    meas1 = m2.meas()
    targetADC_10bit = int(round(meas1 / 0.00375))
    HextargetADC_10bit = hex(targetADC_10bit)
    targetADC_8bit = int(round(meas1 / 0.015))
    HextargetADC_8bit = hex(targetADC_8bit)
    ADC1, ADC2 = a.ADCrslt()
    print(ADC1, ADC2)
    ADCData = (ADC1 * 15)
    ADC2 = ADC2 & 0x03
    ADC = (ADC1 << 2) + ADC2
    LSB_8bit = targetADC_8bit - ADC1
    LSB_10bit = targetADC_10bit - ADC

    localtime = time.asctime(time.localtime(time.time()))
    ws1.cell(i + 2, 1, i)
    ws1.cell(i + 2, 2, localtime)
    ws1.cell(i + 2, 3, meas1)
    ws1.cell(i + 2, 4, HextargetADC_8bit)
    ws1.cell(i + 2, 5, HextargetADC_10bit)
    ws1.cell(i + 2, 6, hex(ADC1))
    ws1.cell(i + 2, 7, hex(ADC))
    ws1.cell(i + 2, 8, LSB_8bit)
    ws1.cell(i + 2, 9, LSB_10bit)
    time.sleep(1)

a.enableFLT(1)
for i in range(n):
    meas1 = m2.meas()
    targetADC = int(round(meas1 / 0.00375))
    HextargetADC = hex(targetADC)
    ADC1, ADC2 = a.ADCrslt()
    print(ADC1, ADC2)
    ADCData = (ADC1 * 15)
    ADC2 = ADC2 & 0x03
    ADC = (ADC1 << 2) + ADC2
    LSB = targetADC - ADC

    localtime = time.asctime(time.localtime(time.time()))
    ws1.cell(i + 2, 12, i)
    ws1.cell(i + 2, 13, localtime)
    ws1.cell(i + 2, 14, meas1)
    ws1.cell(i + 2, 15, HextargetADC_8bit)
    ws1.cell(i + 2, 16, HextargetADC_10bit)
    ws1.cell(i + 2, 17, hex(ADC1))
    ws1.cell(i + 2, 18, hex(ADC))
    ws1.cell(i + 2, 19, LSB_8bit)
    ws1.cell(i + 2, 20, LSB_10bit)
    time.sleep(1)

wb.save(path)
