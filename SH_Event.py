import time
from openpyxl import Workbook
from Mylib.bench_resource import brsc
from ICs import PineHurst
from bit_manipulate import bm
import sys
from dolphin import *

path = './../' + 'SH_Event.xlsx'
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
m1 = brsc.DMM('USB0::0x2A8D::0x1401::MY57215694::0::INSTR')
m2 = brsc.DMM('USB0::0x2A8D::0x1401::MY57218437::0::INSTR')
status_rgt = [0x08, 0x09, 0x0a, 0x0b, 0x33]
clear_rgt = [0x10, 0x11, 0x12, 0x13, 0x14]
mask_rgt = [0x15, 0x16, 0x17, 0x18, 0x19]
statusmap = ['R8[7]VIN_PG', 'R8[6]CTS', 'R8[5]SWA_PG', 'R8[4]SWB_PG', 'R8[3]SWC_PG', 'R8[2]SWD_PG', 'R8[1]VM_OV',
             'R8[0]VIN_OV',
             'R9[7]VIN_HTW', 'R9[6]VB_PG', 'R9[5]LDO1.8PG', 'R9[4]VM2VIN_SO', 'R9[3]SWAHiCurWar', 'R9[2]SWBHiCurWar',
             'R9[1]SWCHiCurWar', 'R9[0]SWDHiCurWar',
             'RA[7]SWAOV', 'RA[6]SWBOV', 'RA[5]SWCOV', 'RA[4]SWDOV', 'RA[3]PECErr', 'RA[2]ParityErr',
             'RA[1]GlobalStatus', 'RA[0]RSV',
             'RB[7]SWACurLim', 'RB[6]SWBCurLim', 'RB[5]SWCCurLim', 'RB[4]SWDCurLim', 'RB[3]SWAUV', 'RB[2]SWBUV',
             'RB[1]SWCUV', 'RB[0]SWDUV',
             'R33[7]TM', 'R33[6]TM', 'R33[5]TM', 'R33[4]VM_PG_SOMO', 'R33[3]VINUV', 'R33[2]LDO1.1PG', 'R33[1]RSV',
             'R33[0]RSV']
maskmap = ['R15[7]VIN_PG', 'R15[6]RSV', 'R15[5]SWAPG', 'R15[4]SWBPG', 'R15[3]SWCPG', 'R15[2]SWDPG', 'R15[1]VM_OV',
           'R15[0]VINOV',
           'R16[7]HTW', 'R16[6]VB_PG', 'R16[5]LDO1.8PG', 'R16[4]VM2VIN_SO', 'R16[3]SWAHICURWAN', 'R16[2]SWBHICURWAN',
           'R16[1]SWCHICURWAN', 'R16[0]SWDHICURWAN',
           'R17[7]SWAOV', 'R17[6]SWBOV', 'R17[5]SWCOV', 'R17[4]SWDOV', 'R17[3]PECERR', 'R17[2]PARITYERR', 'R17[1]RSV',
           'R17[0]RSV',
           'R18[7]SWACURLIM', 'R18[6]SWBCURLIM', 'R18[5]SWCCURLIM', 'R18[4]SWDCURLIM', 'R18[3]SWAUV', 'R18[2]SWBUV',
           'R18[1]SWCUV', 'R18[0]SWDUV',
           'R19[7]RSV', 'R19[6]RSV', 'R19[5]RSV', 'R19[4]VM_PG_SOM', 'R19[3]VB&VIN_UV', 'R19[2]LDO1.0PG', 'R19[1]RSV',
           'R19[0]RSV']


class Piht:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

    def unlock(self):
        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

    def GSI(self, bEnable):
        ret, data = self.dsh.i2c_read_reg(0x1B)
        assert ret == 0
        if bEnable == True:
            data = bm.set_bit(data, 3)
        elif bEnable == False:
            data = bm.clear_bit(data, 3)
        self.dsh.i2c_write_reg(0x1B, data)

    def globalclear(self):
        self.dsh.i2c_write_reg(0x14, 0x01)

    def enable(self, bEnable):
        ret, data = self.dsh.i2c_read_reg(0x32)
        if bEnable:
            data = bm.set_bit(data, 7)
            self.dsh.i2c_write_reg(0x32, data)
        else:
            data = bm.clear_bit(data, 7)
            self.dsh.i2c_write_reg(0x32, data)

    def MaskCtrl(self, Maskctrl):
        ret, data = self.dsh.i2c_read_reg(0x2F)
        if Maskctrl == 0b00:
            data = bm.clear_bit(data, 0, 1)
        elif Maskctrl == 0b01:
            data = bm.clear_bit(data, 1)
            data = bm.set_bit(data, 0)
        elif Maskctrl == 0b10:
            data = bm.set_bit(data, 1)
            data = bm.clear_bit(data, 0)
        elif Maskctrl == 0b11:
            print('Please Check MastCtrl')
            raise
        self.dsh.i2c_write_reg(0x2F, data)


class event_trigger:

    def VIN_PG(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x15, 0, 7)
        elif bMask == 1:
            rgtset(dongle, 0x15, 1, 7)
        a.enable(True)
        time.sleep(1)
        p1.on(4, 2, 2)
        time.sleep(1)
        global name
        name = sys._getframe().f_code.co_name

    def VIN_OV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x15, 0, 0)
        elif bMask == 1:
            rgtset(dongle, 0x15, 1, 0)
        a.enable(True)
        time.sleep(1)
        p1.on(16, 2, 2)
        time.sleep(1)
        global name
        name = sys._getframe().f_code.co_name

    def VIN_MGM_OV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x15, 0, 1)
        elif bMask == 1:
            rgtset(dongle, 0x15, 1, 1)
        a.enable(True)
        time.sleep(1)
        p1.on(4, 0.1, 1)
        time.sleep(1)
        global name
        name = sys._getframe().f_code.co_name

    def VIN_UV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x19, 0, 3)  # mask VIN/Vbias UV
            rgtset(dongle, 0x16, 0, 6)  # mask Vbias PG
        elif bMask == 1:
            rgtset(dongle, 0x19, 1, 3)
            rgtset(dongle, 0x16, 1, 6)
        rgtset(dongle, 0x15, 1, 7)  # mask VIN_PG
        a.enable(True)
        time.sleep(1)
        p1.on(3.5, 2, 2)
        time.sleep(1)
        global name
        name = sys._getframe().f_code.co_name

    def VM_VIN_SOM(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x16, 0, 4)
        elif bMask == 1:
            rgtset(dongle, 0x16, 1, 4)
        a.enable(True)
        time.sleep(1)
        p1.on(2, 2, 1)
        time.sleep(1)
        global name
        name = sys._getframe().f_code.co_name

    def Valid_VM_SOM(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x19, 0, 4)
        elif bMask == 1:
            rgtset(dongle, 0x19, 1, 4)
        rgtset(dongle, 0x16, 1, 4)  # mask VIN_Mgmt to VIN_Bulk Switchover
        a.enable(True)
        time.sleep(1)
        p1.on(2, 2, 1)
        time.sleep(0.3)
        p1.on(3.3, 2, 1)
        time.sleep(1)
        global name
        name = sys._getframe().f_code.co_name

    def VBias_UV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x19, 0, 3)  # mask VIN/Vbias UV
            rgtset(dongle, 0x16, 0, 6)  # mask Vbias PG
        elif bMask == 1:
            rgtset(dongle, 0x19, 1, 3)
            rgtset(dongle, 0x16, 1, 6)
        input('Please Connect Eload CH4 on VBIAS')
        a.enable(True)
        for vol in range(500, 0, -1):
            el1.cv(vol / 100, 5, 4)
            el1.on([4])
            time.sleep(0.005)
            ret, data = dongle.dsh.i2c_read_reg(0x33)
            data = bm.get_bit(data, 's', [3])
            if data == '1':
                print('VBias UV found')
                break
        el1.off([4])
        global name
        name = sys._getframe().f_code.co_name

    def SWAB_CurWarning(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x16, 0, 3)
            rgtset(dongle, 0x16, 0, 2)
        elif bMask == 1:
            rgtset(dongle, 0x16, 1, 3)
            rgtset(dongle, 0x16, 1, 2)
        a.dsh.i2c_write_reg(0x20, 0xFF)  # maxmize current limit threshold
        a.dsh.i2c_write_reg(0x1C, 0x20)  # set SWA current warning 1A
        a.dsh.i2c_write_reg(0x1D, 0x20)  # set SWB current warning 1A
        a.enable(True)
        time.sleep(1)
        for vol in range(110, 30, -1):
            el1.cv(vol / 100, 10, 1)
            el1.on([1])
            time.sleep(0.5)
            ret, data = dongle.dsh.i2c_read_reg(0x0B)
            data = bm.get_bit(data, 's', [7, 6])
            ret, data1 = dongle.dsh.i2c_read_reg(0x09)
            data1 = bm.get_bit(data1, 's', [3, 2])
            if data == '00' and data1 == '11':
                print('VOUTAB Current Warning Found')
                break
        global name
        name = sys._getframe().f_code.co_name

    def SWAB_CurLimit(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x18, 0, 7)
            rgtset(dongle, 0x18, 0, 6)
        elif bMask == 1:
            rgtset(dongle, 0x18, 1, 7)
            rgtset(dongle, 0x18, 1, 6)
        # input('Please Connect DPS CH3 on VOUTAB')
        a.dsh.i2c_write_reg(0x1C, 0xFF)  # maxmize current warning
        a.dsh.i2c_write_reg(0x1D, 0xFF)  # maxmize current warning
        a.enable(True)
        time.sleep(1)
        # p1.SS(1.1, 0.9, 1, 3)
        for vol in range(110, 10, -1):
            # p1.on(vol / 100, 2, 1)
            time.sleep(0.2)
            el1.cv(vol / 100, 10, 1)
            el1.on([1])
            time.sleep(0.5)
            ret, data = dongle.dsh.i2c_read_reg(0x0B)
            data = bm.get_bit(data, 's', [7, 6, 3, 2])
            if data[0] == '1' and data[1] == '1' and data[2] == '0' and data[3] == '0':
                print('VOUTAB Current Limit Found')
                break
        p1.off(3)
        global name
        name = sys._getframe().f_code.co_name

    def SWC_CurWarning(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x16, 0, 1)
        elif bMask == 1:
            rgtset(dongle, 0x16, 1, 1)
        a.dsh.i2c_write_reg(0x20, 0xFF)  # maxmize current limit threshold
        a.dsh.i2c_write_reg(0x1E, 0x20)  # set SWC current warning 1A
        a.enable(True)
        time.sleep(1)
        for vol in range(110, 30, -1):
            el1.cv(vol / 100, 5, 3)
            el1.on([3])
            time.sleep(0.5)
            ret, data = dongle.dsh.i2c_read_reg(0x0B)
            data = bm.get_bit(data, 's', [5])
            ret, data1 = dongle.dsh.i2c_read_reg(0x09)
            data1 = bm.get_bit(data1, 's', [1])
            if data == '0' and data1 == '1':
                print('VOUTC Current Warning Found')
                break
        global name
        name = sys._getframe().f_code.co_name

    def SWC_CurLimit(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x18, 0, 5)
        elif bMask == 1:
            rgtset(dongle, 0x18, 1, 5)
        # input('Please Connect DPS CH3 on VOUTC')
        a.dsh.i2c_write_reg(0x1E, 0xFF)  # maxmize current warning
        rgtset(dongle, 0x26, 1, 2)  # Lowest SWC UV threshold
        a.enable(True)
        time.sleep(1)
        # p1.SS(1.1, 0.9, 1, 3)
        for vol in range(110, 10, -1):
            # p1.on(vol / 100, 2, 1)
            time.sleep(0.2)
            el1.cv(vol / 100, 4.9, 3)
            el1.on([3])
            time.sleep(0.5)
            ret, data = dongle.dsh.i2c_read_reg(0x0B)
            data = bm.get_bit(data, 's', [5, 1])
            # print(data)
            if data[0] == '1' and data[1] == '0':
                print('VOUTC Current Limit Found')
                break
        p1.off(3)
        global name
        name = sys._getframe().f_code.co_name

    def SWD_CurWarning(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x16, 0, 0)
        elif bMask == 1:
            rgtset(dongle, 0x16, 1, 0)
        a.dsh.i2c_write_reg(0x20, 0xFF)  # maxmize current limit threshold
        a.dsh.i2c_write_reg(0x1F, 0x20)  # set SWD current warning 1A
        a.enable(True)
        time.sleep(1)
        for vol in range(180, 30, -1):
            el1.cv(vol / 100, 5, 4)
            el1.on([4])
            time.sleep(0.5)
            ret, data = dongle.dsh.i2c_read_reg(0x0B)
            data = bm.get_bit(data, 's', [4])
            ret, data1 = dongle.dsh.i2c_read_reg(0x09)
            data1 = bm.get_bit(data1, 's', [0])
            if data == '0' and data1 == '1':
                print('VOUTD Current Warning Found')
                break
        global name
        name = sys._getframe().f_code.co_name

    def SWD_CurLimit(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x18, 0, 4)
        elif bMask == 1:
            rgtset(dongle, 0x18, 1, 4)
        # input('Please Connect DPS CH3 on VOUTD')
        a.dsh.i2c_write_reg(0x1F, 0xFF)  # maxmize current warning
        p1.on(1.8, 1, 3)
        a.enable(True)
        time.sleep(1)
        time.sleep(0.3)
        # p1.SS(1.8, 1.7, 1, 3)
        for vol in range(180, 10, -1):
            # p1.on(vol / 100, 2, 1)
            time.sleep(0.2)
            el1.cv(vol / 100, 4.5, 4)
            el1.on([4])
            time.sleep(0.5)
            ret, data = dongle.dsh.i2c_read_reg(0x0B)
            data = bm.get_bit(data, 's', [4, 0])
            # print(data)
            if data[0] == '1' and data[1] == '0':
                print('VOUTD Current Limit Found')
                break
        p1.off(3)
        global name
        name = sys._getframe().f_code.co_name

    def SWAB_PGH(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x15, 0, 5)
        elif bMask == 1:
            rgtset(dongle, 0x15, 1, 5)
        input('Please Connect DPS CH3 on VOUTAB')
        p1.on(1.1, 2, 3)
        a.enable(True)
        time.sleep(1)
        rgtset(dongle, 0x84, 0, 5)  # disable Soft_OV
        p1.SS(1.1, 1.2, 2, 3)
        time.sleep(0.5)
        p1.off(3)
        global name
        name = sys._getframe().f_code.co_name

    def SWC_PGH(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x15, 0, 3)
        elif bMask == 1:
            rgtset(dongle, 0x15, 1, 3)
        input('Please Connect DPS CH3 on VOUTC')
        p1.on(1.1, 1, 3)
        a.enable(True)
        time.sleep(1)
        rgtset(dongle, 0xa3, 0, 5)  # disable Soft_OV
        p1.SS(1.1, 1.2, 1, 3)
        time.sleep(0.5)
        p1.off(3)
        global name
        name = sys._getframe().f_code.co_name

    def SWD_PGH(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x15, 0, 2)
        elif bMask == 1:
            rgtset(dongle, 0x15, 1, 2)
        input('Please Connect DPS CH3 on VOUTD')
        p1.on(1.8, 1, 3)
        a.enable(True)
        time.sleep(1)
        rgtset(dongle, 0xa3, 0, 5)  # disable Soft_OV
        p1.SS(1.8, 1.95, 1, 3)
        time.sleep(0.5)
        p1.off(3)
        global name
        name = sys._getframe().f_code.co_name

    def SWAB_OV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x17, 0, 7)
            rgtset(dongle, 0x17, 0, 6)
        elif bMask == 1:
            rgtset(dongle, 0x17, 1, 7)
            rgtset(dongle, 0x17, 1, 6)
        rgtset(dongle, 0x15, 1, 5)
        rgtset(dongle, 0x15, 1, 4)
        input('Please Connect DPS CH3 on VOUTAB')
        p1.on(1.1, 2, 3)
        a.enable(True)
        time.sleep(1)
        rgtset(dongle, 0x84, 0, 5)  # disable Soft_OV
        p1.SS(1.1, 1.5, 2, 3)
        time.sleep(0.5)
        p1.off(3)
        global name
        name = sys._getframe().f_code.co_name

    def SWC_OV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x17, 0, 5)
        elif bMask == 1:
            rgtset(dongle, 0x17, 1, 5)
        rgtset(dongle, 0x15, 1, 3)
        input('Please Connect DPS CH3 on VOUTC')
        p1.on(1.1, 1, 3)
        a.enable(True)
        time.sleep(1)
        rgtset(dongle, 0xa3, 0, 5)  # disable Soft_OV
        p1.SS(1.1, 1.5, 1, 3)
        time.sleep(0.5)
        p1.off(3)
        global name
        name = sys._getframe().f_code.co_name

    def SWD_OV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x17, 0, 4)
        elif bMask == 1:
            rgtset(dongle, 0x17, 1, 4)
        rgtset(dongle, 0x15, 1, 2)
        input('Please Connect DPS CH3 on VOUTD')
        p1.on(1.8, 1, 3)
        a.enable(True)
        time.sleep(1)
        rgtset(dongle, 0xa3, 0, 5)  # disable Soft_OV
        p1.SS(1.8, 2.1, 1, 3)
        time.sleep(0.5)
        p1.off(3)
        global name
        name = sys._getframe().f_code.co_name

    def SWABUV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x18, 0, 3)
            rgtset(dongle, 0x18, 0, 2)
        elif bMask == 1:
            rgtset(dongle, 0x18, 1, 3)
            rgtset(dongle, 0x18, 1, 2)
        rgtset(dongle, 0x15, 1, 5)  # mask SWA PG
        rgtset(dongle, 0x15, 1, 4)  # mask SWB PG
        rgtset(dongle, 0x18, 1, 7)  # mask SWA Current Limit
        rgtset(dongle, 0x18, 1, 6)  # mask SWB Current Limit
        a.dsh.i2c_write_reg(0x1C, 0xFF)  # maxmize current warning
        a.dsh.i2c_write_reg(0x1D, 0xFF)  # maxmize current warning
        a.enable(True)
        time.sleep(1)
        el1.cc(10, 1)
        el1.on([1])
        time.sleep(1)
        el1.off([1])
        global name
        name = sys._getframe().f_code.co_name

    def SWCUV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x18, 0, 1)
        elif bMask == 1:
            rgtset(dongle, 0x18, 1, 1)
        rgtset(dongle, 0x15, 1, 3)  # mask SWC PG
        rgtset(dongle, 0x18, 1, 5)  # mask SWC Current Limit
        a.dsh.i2c_write_reg(0x1E, 0xFF)  # maxmize current warning
        a.enable(True)
        time.sleep(1)
        el1.cc(5, 3)
        el1.on([3])
        time.sleep(1)
        el1.off([3])
        global name
        name = sys._getframe().f_code.co_name

    def SWDUV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x18, 0, 0)
        elif bMask == 1:
            rgtset(dongle, 0x18, 1, 0)
        rgtset(dongle, 0x15, 1, 2)  # mask SWD PG
        rgtset(dongle, 0x18, 1, 4)  # mask SWD Current Limit
        a.dsh.i2c_write_reg(0x1F, 0xFF)  # maxmize current warning
        a.enable(True)
        time.sleep(1)
        el1.cc(4.5, 4)
        el1.on([4])
        time.sleep(1)
        el1.off([4])
        global name
        name = sys._getframe().f_code.co_name

    def LDO1p8PG(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x16, 0, 5)
        elif bMask == 1:
            rgtset(dongle, 0x16, 1, 5)
        input('Please Connect Eload CH4 on LDO1.8 and DPS on GP20')
        a.enable(True)
        for vol in range(180, 30, -1):
            el1.cv(vol / 100, 5, 4)
            el1.on([4])
            time.sleep(0.005)
            ret, data = dongle.dsh.i2c_read_reg(0x09)
            data = bm.get_bit(data, 's', [5])
            if data == '1':
                print('LDO1.8 PG found')
                break
        el1.off([4])
        global name
        name = sys._getframe().f_code.co_name

    def LDO1p1PG(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x19, 0, 2)
        elif bMask == 1:
            rgtset(dongle, 0x19, 1, 2)
        input('Please Connect Eload CH4 on LDO1.1')
        a.enable(True)
        time.sleep(1)
        el1.cc(0.5, 4)
        el1.on([4])
        time.sleep(1)
        el1.off([4])
        global name
        name = sys._getframe().f_code.co_name

    def HTW(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x16, 0, 7)
        elif bMask == 1:
            rgtset(dongle, 0x16, 1, 7)
        input('Please Heat The Chip')
        a.enable(True)
        while 1:
            rgt = []
            for i in status_rgt:
                ret, data = dongle.dsh.i2c_read_reg(i)
                if ret != 0:
                    rgt.append(data)
                elif ret == 0:
                    rgt.append(hex(data))
            print(rgt)
            if rgt[0] == '0x40':  # or rgt[1] == '0x80':
                print('Critical Temperature Triggered')
                break
        global name
        name = sys._getframe().f_code.co_name


class clear_event:
    def clear_VIN_PG(self, dongle):
        rgtset(dongle, 0x10, 1, 7)  # clear VIN PG

    def clear_VIN_OV(self, dongle):
        rgtset(dongle, 0x10, 1, 0)  # clear VIN OV

    def clear_VIN_MGMT_OV(self, dongle):
        rgtset(dongle, 0x10, 1, 1)  # clear VIN_MGMT OV

    def clear_VM_VIN_SOM(self, dongle):
        rgtset(dongle, 0x11, 1, 4)

    def clear_Valid_VM_SOM(self, dongle):
        rgtset(dongle, 0x14, 1, 4)
        rgtset(dongle, 0x11, 1, 4)

    def clear_VIN_UV(self, dongle):
        rgtset(dongle, 0x14, 1, 3)  # clear VIN UV
        rgtset(dongle, 0x10, 1, 7)  # clear VIN PG
        rgtset(dongle, 0x11, 1, 6)  # clear VBias PG

    def clear_VBIAS_UV(self, dongle):
        rgtset(dongle, 0x14, 1, 3)  # clear VBIAS UV
        rgtset(dongle, 0x11, 1, 6)  # clear VBias PG

    def clear_SWAB_CurLimit(self, dongle):
        rgtset(dongle, 0x13, 1, 7)  # clear SWA Current Limit
        rgtset(dongle, 0x13, 1, 6)  # clear SWB Current Limit

    def clear_SWAB_CurWarning(self, dongle):
        rgtset(dongle, 0x11, 1, 3)  # clear SWA Current Warning
        rgtset(dongle, 0x11, 1, 2)  # clear SWB Current Warning

    def clear_SWC_CurLimit(self, dongle):
        rgtset(dongle, 0x13, 1, 5)  # clear SWB Current Limit

    def clear_SWC_CurWarning(self, dongle):
        rgtset(dongle, 0x11, 1, 1)  # clear SWB Current Warning

    def clear_SWD_CurLimit(self, dongle):
        rgtset(dongle, 0x13, 1, 4)  # clear SWC Current Limit

    def clear_SWD_CurWarning(self, dongle):
        rgtset(dongle, 0x11, 1, 0)  # clear SWC Current Warning

    def clear_SWAB_OV(self, dongle):
        rgtset(dongle, 0x12, 1, 7)  # clear SWA OV
        rgtset(dongle, 0x10, 1, 5)  # clear SWA PG

    def clear_SWC_OV(self, dongle):
        rgtset(dongle, 0x12, 1, 5)  # clear SWB OV
        rgtset(dongle, 0x10, 1, 3)  # clear SWB PG

    def clear_SWD_OV(self, dongle):
        rgtset(dongle, 0x12, 1, 4)  # clear SWC OV
        rgtset(dongle, 0x10, 1, 2)  # clear SWC PG

    def clear_SWAB_PGH(self, dongle):
        rgtset(dongle, 0x10, 1, 5)  # clear SWA PG

    def clear_SWC_PGH(self, dongle):
        rgtset(dongle, 0x10, 1, 3)  # clear SWB PG

    def clear_SWD_PGH(self, dongle):
        rgtset(dongle, 0x10, 1, 2)  # clear SWC PG

    def clear_LDO1p1PG(self, dongle):
        rgtset(dongle, 0x14, 1, 2)  # clear LDO1.1 PG

    def clear_LDO1p8PG(self, dongle):
        rgtset(dongle, 0x11, 1, 5)  # clear LDO1.8 PG

    def clear_SWABUV(self, dongle):
        rgtset(dongle, 0x13, 1, 3)  # clear SWAUV
        rgtset(dongle, 0x13, 1, 7)  # clear SWA Current Limit
        rgtset(dongle, 0x10, 1, 5)  # clear SWA PG
        rgtset(dongle, 0x13, 1, 2)  # clear SWBUV
        rgtset(dongle, 0x13, 1, 6)  # clear SWB Current Limit
        rgtset(dongle, 0x10, 1, 2)  # clear SWB PG

    def clear_SWCUV(self, dongle):
        rgtset(dongle, 0x13, 1, 1)  # clear SWCUV
        rgtset(dongle, 0x13, 1, 5)  # clear SWC Current Limit
        rgtset(dongle, 0x10, 1, 3)  # clear SWC PG

    def clear_SWDUV(self, dongle):
        rgtset(dongle, 0x13, 1, 0)  # clear SWCUV
        rgtset(dongle, 0x13, 1, 4)  # clear SWC Current Limit
        rgtset(dongle, 0x10, 1, 2)  # clear SWC PG

    def clear_HTW(self, dongle):
        rgtset(dongle, 0x11, 1, 7)  # clear HTW


def maskcheck(dongle):
    rgt = []
    maskindex = []
    for i in mask_rgt:
        ret, data = dongle.dsh.i2c_read_reg(i)
        rgt.append(hex(data))
        index = mask_rgt.index(i)
        str1 = str(bin(data))[2:10].zfill(8)
        for index1, s in enumerate(str1):
            if s == '1':
                maskindex.append(maskmap[index1 + index * 8])
    print(maskindex)


def statuscheck(dongle):
    rgt = []
    statusindex = ''
    for i in status_rgt:
        ret, data = dongle.dsh.i2c_read_reg(i)
        if ret != 0:
            rgt.append(data)
        elif ret == 0:
            rgt.append(hex(data))
            index = status_rgt.index(i)
            str1 = str(bin(data))[2:10].zfill(8)
            for index1, s in enumerate(str1):
                if s == '1':
                    statusindex = statusindex + ';' + str(statusmap[index1 + index * 8])

        VOUTA = el1.volmeas(1)
        VOUTB = el1.volmeas(2)
        VOUTC = el1.volmeas(3)
        VOUTD = el1.volmeas(4)
        Iin = p1.curmeas(2)
        PG = m1.meas()
        GSI = m2.meas()

        if VOUTA < 0.3 or VOUTB < 0.3 or VOUTC < 0.3 or VOUTD < 0.3 or Iin < 0.001:
            VR = 'disabled'
        else:
            VR = 'enabled'

        if PG > 1.6:
            PG = 'H'
        elif PG < 0.3:
            PG = 'L'
        else:
            PG = 'Please Check'

        if GSI > 1.6:
            GSI = 'H'
        elif GSI < 0.3:
            GSI = 'L'
        else:
            GSI = 'Please Check'

    return rgt, VR, PG, GSI, statusindex


def rgtset(dongle, rgt, set, *n):
    ret, data = dongle.dsh.i2c_read_reg(rgt)
    if set == 1:
        data = bm.set_bit(data, *n)
    elif set == 0:
        data = bm.clear_bit(data, *n)
    dongle.dsh.i2c_write_reg(rgt, data)


wb = Workbook()
ws1 = wb.active

ws1.cell(row=1, column=1, value='Item')
ws1.cell(row=1, column=2, value='Status Register')
ws1.cell(row=1, column=3, value='VR')
ws1.cell(row=1, column=4, value='PG')
ws1.cell(row=1, column=5, value='GSI')
ws1.cell(row=1, column=6, value='Status Value')
ws1.cell(row=1, column=7, value='Clear Flag')
ws1.cell(row=1, column=8, value='VR_Cleared')
ws1.cell(row=1, column=9, value='PG_Cleared')
ws1.cell(row=1, column=10, value='GSI_Cleared')

a = Piht()
t = event_trigger()
c = clear_event()
mask = 1

for i in range(19, 20, 1):
    p1.off(1)
    p1.off(2)
    time.sleep(1)
    p1.on(3.3, 1, 1)
    p1.on(12, 2, 2)
    time.sleep(1)

    a.unlock()
    a.GSI(True)
    a.MaskCtrl(0b10)

    if i == 0:
        t.VIN_PG(a, mask)

    elif i == 1:
        t.VIN_OV(a, mask)

    elif i == 2:
        t.VIN_MGM_OV(a, mask)

    elif i == 3:
        t.VIN_UV(a, mask)

    elif i == 4:
        t.VBias_UV(a, mask)

    elif i == 5:
        t.SWAB_CurLimit(a, mask)

    elif i == 6:
        t.SWAB_CurWarning(a, mask)

    elif i == 7:
        t.SWC_CurLimit(a, mask)

    elif i == 8:
        t.SWC_CurWarning(a, mask)

    elif i == 9:
        t.SWD_CurLimit(a, mask)

    elif i == 10:
        t.SWD_CurWarning(a, mask)

    elif i == 11:
        t.SWAB_OV(a, mask)

    elif i == 12:
        t.SWAB_PGH(a, mask)

    elif i == 13:
        t.SWC_OV(a, mask)

    elif i == 14:
        t.SWC_PGH(a, mask)

    elif i == 15:
        t.SWD_OV(a, mask)

    elif i == 16:
        t.SWD_PGH(a, mask)

    elif i == 17:
        t.SWABUV(a, mask)

    elif i == 18:
        t.SWCUV(a, mask)

    elif i == 19:
        t.SWDUV(a, mask)

    elif i == 20:
        t.LDO1p1PG(a, mask)

    elif i == 21:
        t.LDO1p8PG(a, mask)

    elif i == 22:
        t.VM_VIN_SOM(a, mask)

    elif i == 23:
        t.Valid_VM_SOM(a, mask)

    elif i == 24:
        t.HTW(a, mask)

    status, VR, PG, GSI, statusvalue = statuscheck(a)
    print('{} status: {}; VR: {}; PG: {}; GSI: {};statusvalue: {}'.format(name, status, VR, PG, GSI, statusvalue))

    if i != 22 and i != 23:
        p1.on(3.3, 0.1, 1)
    p1.on(12, 2, 2)
    p1.off(3)
    el1.off([1, 2, 3, 4])
    time.sleep(1)
    flag = 0

    if i == 0:
        c.clear_VIN_PG(a)

    elif i == 1:
        c.clear_VIN_OV(a)

    elif i == 2:
        c.clear_VIN_MGMT_OV(a)

    elif i == 3:
        c.clear_VIN_UV(a)

    elif i == 4:
        c.clear_VBIAS_UV(a)

    elif i == 5:
        c.clear_SWAB_CurLimit(a)

    elif i == 6:
        c.clear_SWAB_CurWarning(a)

    elif i == 7:
        c.clear_SWC_CurLimit(a)

    elif i == 8:
        c.clear_SWC_CurWarning(a)

    elif i == 9:
        c.clear_SWD_CurLimit(a)

    elif i == 10:
        c.clear_SWD_CurWarning(a)

    elif i == 11:
        c.clear_SWAB_OV(a)

    elif i == 12:
        c.clear_SWAB_PGH(a)

    elif i == 13:
        c.clear_SWC_OV(a)

    elif i == 14:
        c.clear_SWC_PGH(a)

    elif i == 15:
        c.clear_SWD_OV(a)

    elif i == 16:
        c.clear_SWD_PGH(a)

    elif i == 17:
        c.clear_SWABUV(a)

    elif i == 18:
        c.clear_SWCUV(a)

    elif i == 19:
        c.clear_SWDUV(a)

    elif i == 20:
        c.clear_LDO1p1PG(a)

    elif i == 21:
        c.clear_LDO1p8PG(a)

    elif i == 22:
        c.clear_VM_VIN_SOM(a)

    elif i == 23:
        c.clear_Valid_VM_SOM(a)

    elif i == 24:
        c.clear_HTW(a)

    statusC, VRC, PGC, GSIC, statusvalueC = statuscheck(a)
    print('{} statusC: {}; VRC: {}; PGC: {}; GSIC: {};statusvalueC: {}'.format(name, statusC, VRC, PGC, GSIC,
                                                                               statusvalueC))
    if statusvalueC == '':
        flag = 0
    else:
        flag = statusvalueC

    ws1.cell(row=2 + i, column=1, value=name)
    ws1.cell(row=2 + i, column=2, value=(str(status[0]) + ',' + str(status[1])) + ',' + str(status[2]) + ',' + str(
        status[3]) + ',' + str(status[4]))
    ws1.cell(row=2 + i, column=3, value=VR)
    ws1.cell(row=2 + i, column=4, value=PG)
    ws1.cell(row=2 + i, column=5, value=GSI)
    ws1.cell(row=2 + i, column=6, value=statusvalue)
    ws1.cell(row=2 + i, column=7, value=flag)
    ws1.cell(row=2 + i, column=8, value=VRC)
    ws1.cell(row=2 + i, column=9, value=PGC)
    ws1.cell(row=2 + i, column=10, value=GSIC)

    wb.save(path)
