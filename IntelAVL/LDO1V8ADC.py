import time
from ICs import PineHurst
from openpyxl import Workbook
from bit_manipulate import bm
from Mylib.bench_resource import brsc


class PHCurMeas:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

    def enable(self):
        # enable
        ret, data = self.dsh.i2c_read_reg(0x2F)
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)
        self.dsh.i2c_write_reg(0x32, 0x80)

    def LDO1p8Select(self, vol):
        if vol == 0:
            rgt = 0x02
        elif vol == 1:
            rgt = 0x42
        elif vol == 2:
            rgt = 0x82
        elif vol == 3:
            rgt = 0xc2
        self.dsh.i2c_write_reg(0x2b, rgt)
        print(self.dsh.i2c_read_reg(0x2b))

    def ADCMeas(self):
        self.dsh.i2c_write_reg(0x30, 0xc0)
        ret, data = self.dsh.i2c_read_reg(0x31)
        return data


path = './LDO1P8ADC_' + '.xlsx'

wb = Workbook()
ws1 = wb.active
ws1.cell(row=1, column=1, value='Meas')
ws1.cell(row=1, column=2, value='ADC')

p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
m1 = brsc.DMM('USB0::0x2A8D::0x1401::MY57215694::0::INSTR')
p1.on(5, 2, 1)
p = PHCurMeas()
# p.enable()
time.sleep(1)

for i in range(4):
    time.sleep(0.5)
    p.LDO1p8Select(i)
    meas = m1.meas()
    ADC = p.ADCMeas()
    ws1.cell(row=2 + i, column=1, value=1000*meas)
    ws1.cell(row=2 + i, column=2, value=hex(ADC)[2:])

wb.save(path)
