import time
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from Mylib.bench_resource import brsc
from ICs import PineHurst
from bit_manipulate import bm
from openpyxl.styles import colors, Font


class Piht:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

    def enable(self, bEnable):
        if bEnable:
            self.dsh.i2c_write_reg(0x32, 0x80)
        else:
            self.dsh.i2c_write_reg(0x32, 0x00)

    def freq(self, ch, freq):
        if ch == 'a':
            ret, data = self.dsh.i2c_read_reg(0x29)
            if freq == 1000:
                data = bm.set_bit(data, 4)
                data = bm.clear_bit(data, 5)
                self.dsh.i2c_write_reg(0x29, data)
            elif freq == 750:
                data = bm.clear_bit(data, 5, 4)
                self.dsh.i2c_write_reg(0x29, data)
        elif ch == 'b':
            ret, data = self.dsh.i2c_read_reg(0x2A)
            if freq == 1000:
                data = bm.set_bit(data, 4)
                data = bm.clear_bit(data, 5)
                self.dsh.i2c_write_reg(0x2A, data)
            elif freq == 750:
                data = bm.clear_bit(data, 5, 4)
                self.dsh.i2c_write_reg(0x2A, data)
        elif ch == 'c':
            ret, data = self.dsh.i2c_read_reg(0x2A)
            if freq == 750:
                data = bm.clear_bit(data, 1, 0)
                self.dsh.i2c_write_reg(0x2A, data)
            elif freq == 1000:
                data = bm.set_bit(data, 0)
                data = bm.clear_bit(data, 1)
                self.dsh.i2c_write_reg(0x2A, data)
        print(self.dsh.i2c_read_reg(0x29), self.dsh.i2c_read_reg(0x2A))


path = './' + 'EFF.xlsx'
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
m1 = brsc.DMM('USB0::0x2A8D::0x1401::MY57215694::0::INSTR')
m2 = brsc.DMM('USB0::0x2A8D::0x1401::MY57218437::0::INSTR')
p1.on(5, 2, 1)

wb = Workbook()
ws1 = wb.active

Rail = 'b'
a = Piht()
a.freq(Rail, 750)

if Rail == 'a':
    a.dsh.i2c_write_reg(0x2f, 0x42)  # RAILA
    channel = 1
    load = [0.1, 0.2, 0.4, 0.6, 0.8, 1, 1.2, 1.4, 1.6, 1.8, 2, 2.2, 2.4, 2.6, 2.8, 3, 3.4, 3.8, 4]
elif Rail == 'b':
    a.dsh.i2c_write_reg(0x2f, 0x12)  # RAILB
    channel = 2
    load = [0.1, 0.2, 0.4, 0.6, 0.8, 1, 1.2, 1.4, 1.6, 1.8, 2, 2.2, 2.4, 2.6, 2.8, 3, 3.4, 3.8, 4]
elif Rail == 'c':
    a.dsh.i2c_write_reg(0x2f, 0x0A)  # RAILC
    channel = 3
    load = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1]

step1 = 0.2
step2 = 0.4
group = 0
maxcell = 0
i = 0

while True:
    print(i, load[i])

    el1.cc(load[i], channel)
    el1.on([channel])
    time.sleep(0.5)

    iout = el1.curmeas(channel)
    v1 = m1.meas()
    v2 = m2.meas()
    Iin_p = p1.curmeas(1)

    power_in = v1 * Iin_p
    power_out = v2 * iout
    Eff = 100 * power_out / power_in
    if maxcell < Eff:
        maxcell = Eff
        red_c = 5 + 8 * group
        red_r = i + 2

    ws1.cell(row=i + 2, column=1 + 8 * group, value=v1)
    ws1.cell(row=i + 2, column=2 + 8 * group, value=Iin_p)
    ws1.cell(row=i + 2, column=3 + 8 * group, value=v2)
    ws1.cell(row=i + 2, column=4 + 8 * group, value=iout)
    ws1.cell(row=i + 2, column=5 + 8 * group, value=Eff)

    ws1.cell(row=1, column=1 + 8 * group, value='Vin_Bulk')
    ws1.cell(row=1, column=2 + 8 * group, value='Iin_Bulk')
    ws1.cell(row=1, column=3 + 8 * group, value='VOUT')
    ws1.cell(row=1, column=4 + 8 * group, value='IOUT')
    ws1.cell(row=1, column=5 + 8 * group, value='Efficiency')

    cnt = len(load)
    if i < cnt - 1:
        i += 1
    else:
        break

ft = Font(color=colors.RED, size=12)
c = ws1.cell(row=red_r, column=red_c)
c.font = ft

wb.save(path)

el1.off([channel])
