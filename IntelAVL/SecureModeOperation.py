# -*- coding: utf-8 -*-

import sys
import time
from bit_manipulate import bm
from openpyxl import Workbook
from ICs import DolphinSH
from Mylib.bench_resource import brsc


class SHSecureMode:
    def __init__(self):
        self.dsh = DolphinSH()
        ret = self.dsh.connect()
        if ret is False:
            print("Failed to connect")
            sys.exit(-1)

        # self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

    def unlockDIMM(self, bEnable):
        if bEnable == 1:
            # # Un-lock DIMM vendor region(0x40~0x6f)
            self.dsh.i2c_write_reg(0x37, 0x73)
            self.dsh.i2c_write_reg(0x38, 0x94)
            self.dsh.i2c_write_reg(0x39, 0x40)

    def unlockIDT(self, bEnable):
        if bEnable == 1:
            # # Un-lock PMIC vendor region(0x70~0xff)
            self.dsh.i2c_write_reg(0x37, 0x79)
            self.dsh.i2c_write_reg(0x38, 0xbe)
            self.dsh.i2c_write_reg(0x39, 0x10)

    def securemode(self, bEnable):
        ret, data = self.dsh.i2c_read_reg(0x2F)
        if bEnable == 0:
            data = bm.set_bit(data, 2)
        elif bEnable == 1:
            data = bm.clear_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)
        ret, data = self.dsh.i2c_read_reg(0x2F)
        print('0x2F', hex(data))

    def enable(self, bEnable):
        if bEnable:
            self.dsh.i2c_write_reg(0x32, 0x80)
        else:
            self.dsh.i2c_write_reg(0x32, 0x00)


f1 = open('D:\DDR_PMIC\Pinehurst\Intel_AVL\SecureModeOperation.txt', 'r+')

print(f1.name)
rgtlist, datalist = [], []
for line in f1.readlines():
    if line.count('>') > 0 and line.count('<') > 0:
        start = line.find('>') + 1
        end = line[1:].find('<') + 1
        if end > start:
            str1 = line[start:end]
            if len(str1) > 2:
                rgt = str1[0:2]
                data = str1[3:5]
                rgtlist.append(rgt)
                datalist.append(data)
f1.close()

print(rgtlist)
print(datalist)

path = './' + 'SecureModeOperation.xlsx'
wb = Workbook()
ws1 = wb.active
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')

p1.off(2)
time.sleep(1)
p1.on(5, 2, 1)
time.sleep(1)
a = SHSecureMode()
a.securemode(1)
a.unlockDIMM(0)
a.unlockIDT(0)
a.enable(1)
time.sleep(1)

for i in rgtlist:
    index = rgtlist.index(i)
    ret, initial = a.dsh.i2c_read_reg(int(i, 16))
    a.dsh.i2c_write_reg(int(i, 16), int(datalist[index], 16))
    ret, final = a.dsh.i2c_read_reg(int(i, 16))
    print(i, hex(initial), hex(final))
    ws1.cell(row=index + 1, column=1, value=i)
    ws1.cell(row=index + 1, column=2, value=hex(initial))
    ws1.cell(row=index + 1, column=3, value=hex(final))
    ws1.cell(row=index + 1, column=4, value=hex(int(datalist[index], 16)))

wb.save(path)
