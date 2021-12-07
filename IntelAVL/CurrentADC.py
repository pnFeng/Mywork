import time
from ICs import PineHurst
from openpyxl import Workbook
from bit_manipulate import bm
from Mylib.bench_resource import brsc


class PHCurMeas:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

    def enable(self):
        # enable
        ret, data = self.dsh.i2c_read_reg(0x2F)
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)
        self.dsh.i2c_write_reg(0x32, 0x80)

    def CurMeas(self, rail):
        sum = 0
        self.dsh.i2c_write_reg(0x1B, 0x05)  # Select Current Meas
        if rail == 'a':
            rgt = 0x0C
        elif rail == 'b':
            rgt = 0x0E
        elif rail == 'c':
            rgt = 0x0F
        ret, data = self.dsh.i2c_read_reg(rgt)
        return data

    def PowerMeas(self, rail):
        sum = 0
        self.dsh.i2c_write_reg(0x1B, 0x45)  # Select Power Meas
        if rail == 'a':
            rgt = 0x0C
        elif rail == 'b':
            rgt = 0x0E
        elif rail == 'c':
            rgt = 0x0F
        ret, data = self.dsh.i2c_read_reg(rgt)
        return data


Rail = 'c'
el_ch = 3
path = './CurADC_' + Rail + '.xlsx'
step = 0.5

wb = Workbook()
ws1 = wb.active
ws1.cell(row=1, column=1, value='load')

el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
s1 = brsc.Scope('USB0::0x05FF::0x1023::3808N60392::INSTR')
m2 = brsc.DMM('USB0::0x2A8D::0x1401::MY57218437::0::INSTR')
p1.on(5, 2, 1)
p = PHCurMeas()
p.enable()

load = [0, 0.1, 0.5, 1]
time.sleep(1)

for i in load:
    el1.cc(i, el_ch)
    el1.on([el_ch])
    time.sleep(0.5)
    load1 = el1.curmeas([el_ch])
    time.sleep(1)

    for j in range(10):
        ws1.cell(row=1, column=2 + j * 2, value='ADC_6bit_Current_' + str(j))
        ws1.cell(row=1, column=3 + j * 2, value='Probe Meas')

        vout = m2.meas()
        s1.scope.write('vbs app.measure.clearsweeps')
        time.sleep(1)
        Reading = p.CurMeas(Rail)
        c1 = float(s1.scope.query('vbs? return = app.measure.p1.mean.result.value'))
        if j == 0:
            flag = 'nochange'
            last = Reading
        elif last != Reading:
            flag = 'change'
        r = load.index(i)
        ws1.cell(row=r + 2, column=1, value=i)
        ws1.cell(row=r + 2, column=2 + j * 2, value=hex(Reading)[2:])
        ws1.cell(row=r + 2, column=3 + j * 2, value=c1)
        ws1.cell(row=r + 2, column=22, value=vout)
        ws1.cell(row=r + 2, column=23, value=flag)
        print(i, j)
        time.sleep(0.2)

time.sleep(1)
el1.off([1])
wb.save(path)
