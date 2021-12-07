import time
from ICs import PineHurst
from openpyxl import Workbook
from bit_manipulate import bm
from Mylib.bench_resource import brsc


class PHCurMeas:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

    def enable(self):
        # enable
        ret, data = self.dsh.i2c_read_reg(0x2F)
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)
        self.dsh.i2c_write_reg(0x32, 0x80)

    def CurMeas(self, rail):
        sum = 0
        if rail == 'a':
            rgt = 0x0C
        elif rail == 'b':
            rgt = 0x0E
        elif rail == 'c':
            rgt = 0x0F
        ret, data = self.dsh.i2c_read_reg(rgt)
        return data


Rail = 'c'
el_ch = 3
path = './CurrentADC_' + Rail + '.xlsx'
step = 0.125

wb = Workbook()
ws1 = wb.active
ws1.cell(row=1, column=1, value='load')
ws1.cell(row=1, column=2, value='ADC_6bit_Current')
ws1.cell(row=1, column=3, value='ADC_6bit_Current')
ws1.cell(row=1, column=4, value='ADC_6bit_Current')
ws1.cell(row=1, column=5, value='ADC_6bit_Current')
ws1.cell(row=1, column=6, value='ADC_6bit_Current')
ws1.cell(row=1, column=7, value='ADC_6bit_Current')
ws1.cell(row=1, column=8, value='ADC_6bit_Current')
ws1.cell(row=1, column=9, value='ADC_6bit_Current')
ws1.cell(row=1, column=10, value='ADC_6bit_Current')

el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
p1.on(5, 2, 1)
p = PHCurMeas()
p.enable()
time.sleep(1)

for i in range(9):
    if i == 0:
        load = 0
    else:
        load += step
    el1.cc(load, el_ch)
    el1.on([el_ch])
    time.sleep(0.5)
    load1 = el1.curmeas([el_ch])
    time.sleep(1)

    for j in range(10):
        Reading = p.CurMeas(Rail)
        if j == 0:
            flag = 'nochange'
            last = Reading
        elif last != Reading:
            flag = 'change'
        ws1.cell(row=i + 2, column=1, value=load)
        ws1.cell(row=i + 2, column=2 + j, value=hex(Reading)[2:])
        ws1.cell(row=i + 2, column=12, value=flag)
        print(i, j)
        time.sleep(0.2)

time.sleep(1)
el1.off([1])
wb.save(path)
