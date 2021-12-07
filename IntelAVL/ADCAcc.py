# -*- coding: utf-8 -*-

import time

from bit_manipulate import bm
from openpyxl import Workbook

from ICs import PineHurst
from Mylib.bench_resource import brsc

p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
m2 = brsc.DMM('USB0::0x2A8D::0x1401::MY57218437::0::INSTR')
el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')


class SHVID:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(1000)

        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

        # Diable OVP
        self.dsh.i2c_write_reg(0x56, 0xF0)

    def enable(self):
        # enable
        ret, data = self.dsh.i2c_read_reg(0x2F)
        data = data | 0x04
        self.dsh.i2c_write_reg(0x2F, data)
        self.dsh.i2c_write_reg(0x32, 0x80)
        self.dsh.i2c_write_reg(0x6B, data)

    def VIDSetting(self, ch, VID, cre, freq):
        # self.dsh.i2c_write_reg(0x29, 0xC0)  # Forced SWA CCM
        # self.dsh.i2c_write_reg(0x2A, 0xCC)  # Forced SWB/C CCM
        # self.value = VID << 1 | cre
        if ch == 'a':
            self.dsh.i2c_write_reg(0x21, VID)
            ret, data = self.dsh.i2c_read_reg(0x29)
            if freq == 1000:
                data = bm.set_bit(data, 4)
                data = bm.clear_bit(data, 5)
            elif freq == 750:
                data = bm.clear_bit(data, 5, 4)
            self.dsh.i2c_write_reg(0x29, data)
        elif ch == 'b':
            self.dsh.i2c_write_reg(0x25, VID)
            ret, data = self.dsh.i2c_read_reg(0x2A)
            if freq == 1000:
                data = bm.set_bit(data, 4)
                data = bm.clear_bit(data, 5)
            elif freq == 750:
                data = bm.clear_bit(data, 5, 4)
            self.dsh.i2c_write_reg(0x2A, data)
        elif ch == 'c':
            self.dsh.i2c_write_reg(0x27, VID)
            ret, data = self.dsh.i2c_read_reg(0x2A)
            if freq == 750:
                data = bm.clear_bit(data, 1, 0)
            elif freq == 1000:
                data = bm.set_bit(data, 0)
                data = bm.clear_bit(data, 1)
            self.dsh.i2c_write_reg(0x2A, data)

    def ADCConfig(self, ch):
        if ch == 'a':
            self.dsh.i2c_write_reg(0x30, 0x80)
        elif ch == 'b':
            self.dsh.i2c_write_reg(0x30, 0x90)
        elif ch == 'c':
            self.dsh.i2c_write_reg(0x30, 0x98)

    def ADCrslt(self):
        ret, data1 = self.dsh.i2c_read_reg(0x31)
        ret, data2 = self.dsh.i2c_read_reg(0x72)
        return data1, data2

    def enableFLT(self, enable):
        if enable == 1:
            ret, data = self.dsh.i2c_read_reg(0xc7)
            data = data | 0x40
            self.dsh.i2c_write_reg(0xc7, data)
            ret, data = self.dsh.i2c_read_reg(0xc7)
        elif enable == 0:
            ret, data = self.dsh.i2c_read_reg(0xc7)
            data = data & 0xBF
            self.dsh.i2c_write_reg(0xc7, data)
            ret, data = self.dsh.i2c_read_reg(0xc7)


ch = 'a'
channel = 1

path = './' + 'ADCAcc.xlsx'
wb = Workbook()
ws1 = wb.active

p1.on(5, 2, 1)

time.sleep(1)
a = SHVID()
a.ADCConfig(ch)
a.enable()
a.enableFLT(1)
time.sleep(1)

vol = [5, 5.5, 4.25]
for j in vol:
    p1.on(j, 2, 1)
    localtime = time.asctime(time.localtime(time.time()))
    ws1.cell(1, 1, 'Condition')
    ws1.cell(1, 2, 'Time')
    ws1.cell(1, 3, 'Multimeter VOUT(V)')
    ws1.cell(1, 4, 'ADC')

    # 750k Load=0A
    n = [0x00, 0x78, 0xFE]
    for i in n:
        print(j, hex(i))
        a.VIDSetting(ch, i, 0, 1000)
        time.sleep(0.5)
        ret, data = a.dsh.i2c_read_reg(0x21)
        meas1 = m2.meas()

        Sum = 0
        for k in range(10):
            ADC1, ADC2 = a.ADCrslt()
            ADCData = (ADC1 * 15)
            Sum += ADC1
            time.sleep(0.05)
        ADCavg = int(round(Sum / 10, 0))

        r = n.index(i)
        r1 = vol.index(j)
        ws1.cell(r + 2 + 13 * r1, 1, str(j) + hex(i) + ' load=0' + ' 750k')
        ws1.cell(r + 2 + 13 * r1, 2, localtime)
        ws1.cell(r + 2 + 13 * r1, 3, 1000 * meas1)
        ws1.cell(r + 2 + 13 * r1, 4, hex(ADCavg)[2:])

    # 750k Load=1A
    el1.cc(1, channel)
    el1.on([channel])
    n = [0x00, 0x78, 0xFE]
    for i in n:
        print(j, hex(i))
        a.VIDSetting(ch, i, 0, 750)
        time.sleep(0.5)
        ret, data = a.dsh.i2c_read_reg(0x21)
        meas1 = m2.meas()

        Sum = 0
        for k in range(10):
            ADC1, ADC2 = a.ADCrslt()
            ADCData = (ADC1 * 15)
            Sum += ADC1
            time.sleep(0.05)
        ADCavg = int(round(Sum / 10, 0))

        r = n.index(i)
        r1 = vol.index(j)
        ws1.cell(r + 5 + 13 * r1, 1, str(j) + hex(i) + 'load=1A' + '750k')
        ws1.cell(r + 5 + 13 * r1, 2, localtime)
        ws1.cell(r + 5 + 13 * r1, 3, 1000 * meas1)
        ws1.cell(r + 5 + 13 * r1, 4, hex(ADCavg)[2:])

    # 1000k Load=0A
    el1.off([channel])
    n = [0x00, 0x78, 0xFE]
    for i in n:
        print(j, hex(i))
        a.VIDSetting(ch, i, 0, 1000)
        time.sleep(0.5)
        ret, data = a.dsh.i2c_read_reg(0x21)
        meas1 = m2.meas()

        Sum = 0
        for k in range(10):
            ADC1, ADC2 = a.ADCrslt()
            ADCData = (ADC1 * 15)
            Sum += ADC1
            time.sleep(0.05)
        ADCavg = int(round(Sum / 10, 0))

        r = n.index(i)
        r1 = vol.index(j)
        ws1.cell(r + 8 + 13 * r1, 1, str(j) + hex(i) + 'load=0' + '1000k')
        ws1.cell(r + 8 + 13 * r1, 2, localtime)
        ws1.cell(r + 8 + 13 * r1, 3, 1000 * meas1)
        ws1.cell(r + 8 + 13 * r1, 4, hex(ADCavg)[2:])

    # 1000k Load=1A
    el1.cc(1, channel)
    el1.on([channel])
    n = [0x00, 0x78, 0xFE]
    for i in n:
        print(j, hex(i))
        a.VIDSetting(ch, i, 0, 1000)
        time.sleep(0.5)
        ret, data = a.dsh.i2c_read_reg(0x21)
        meas1 = m2.meas()

        Sum = 0
        for k in range(10):
            ADC1, ADC2 = a.ADCrslt()
            ADCData = (ADC1 * 15)
            Sum += ADC1
            time.sleep(0.05)
        ADCavg = int(round(Sum / 10, 0))

        r = n.index(i)
        r1 = vol.index(j)
        ws1.cell(r + 11 + 13 * r1, 1, str(j) + hex(i) + 'load=1A' + '1000k')
        ws1.cell(r + 11 + 13 * r1, 2, localtime)
        ws1.cell(r + 11 + 13 * r1, 3, 1000 * meas1)
        ws1.cell(r + 11 + 13 * r1, 4, hex(ADCavg)[2:])

el1.off([channel])
wb.save(path)
