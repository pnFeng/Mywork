# -*- coding: utf-8 -*-

import time
from bit_manipulate import bm
from openpyxl import Workbook
from Mylib.bench_resource import brsc
from ICs import PineHurst

p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
s1 = brsc.Scope('USB0::0x05FF::0x1023::3808N60392::INSTR')


class SHVID:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(1000)

        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

    def enable(self):
        # enable
        ret, data = self.dsh.i2c_read_reg(0x2F)
        print('data', data, 'ret', ret)
        data = data | 0x04
        self.dsh.i2c_write_reg(0x2F, data)
        self.dsh.i2c_write_reg(0x32, 0x80)
        self.dsh.i2c_write_reg(0x6B, data)

    def Freq(self, freq):
        # SWA
        ret, data = self.dsh.i2c_read_reg(0x29)
        if freq == 1000:
            data = bm.set_bit(data, 4)
            data = bm.clear_bit(data, 5)
        elif freq == 750:
            data = bm.clear_bit(data, 5, 4)
        self.dsh.i2c_write_reg(0x29, data)

        # SWB
        ret, data = self.dsh.i2c_read_reg(0x2A)
        if freq == 1000:
            data = bm.set_bit(data, 4)
            data = bm.clear_bit(data, 5)
        elif freq == 750:
            data = bm.clear_bit(data, 5, 4)
        self.dsh.i2c_write_reg(0x2A, data)

        # SWC
        ret, data = self.dsh.i2c_read_reg(0x2A)
        if freq == 750:
            data = bm.clear_bit(data, 1, 0)
        elif freq == 1000:
            data = bm.set_bit(data, 0)
            data = bm.clear_bit(data, 1)
        self.dsh.i2c_write_reg(0x2A, data)


def crosscheck(rail, load):
    if rail == 'a':
        if load != 0:
            el1.cc(4, 1)
            el1.on([1])
            time.sleep(0.5)
        s1.scope.write('vbs app.measure.clearsweeps')
        time.sleep(3)
        va_max = float(s1.scope.query('vbs? return = app.measure.p1.max.result.value'))
        va_min = float(s1.scope.query('vbs? return = app.measure.p2.min.result.value'))
        va_mean = float(s1.scope.query('vbs? return = app.measure.p3.mean.result.value'))
        el1.off([1, 2, 3])
        return va_mean, va_max, va_min
    elif rail == 'b':
        if load != 0:
            el1.cc(4, 2)
            el1.on([2])
            time.sleep(0.5)
        s1.scope.write('vbs app.measure.clearsweeps')
        time.sleep(3)
        vb_max = float(s1.scope.query('vbs? return = app.measure.p4.max.result.value'))
        vb_min = float(s1.scope.query('vbs? return = app.measure.p5.min.result.value'))
        vb_mean = float(s1.scope.query('vbs? return = app.measure.p6.mean.result.value'))
        el1.off([1, 2, 3])
        return vb_mean, vb_max, vb_min
    elif rail == 'c':
        if load != 0:
            el1.cc(1, 3)
            el1.on([3])
            time.sleep(0.5)
        s1.scope.write('vbs app.measure.clearsweeps')
        time.sleep(3)
        vc_max = float(s1.scope.query('vbs? return = app.measure.p7.max.result.value'))
        vc_min = float(s1.scope.query('vbs? return = app.measure.p8.min.result.value'))
        vc_mean = float(s1.scope.query('vbs? return = app.measure.p9.mean.result.value'))
        el1.off([1, 2, 3])
        return vc_mean, vc_max, vc_min


path = './' + 'CrossRegulation.xlsx'
wb = Workbook()
ws1 = wb.active

p1.on(5.5, 4, 1)
time.sleep(1)
a = SHVID()
a.Freq(750)
a.enable()
time.sleep(1)

ws1.cell(1, 1, 'Mean')
ws1.cell(1, 2, 'Max')
ws1.cell(1, 3, 'Min')

for i in range(12):

    if i == 0:
        input('Please add aggressor a')
        mean, max, min = crosscheck('b', 0)

    elif i == 1:
        mean, max, min = crosscheck('c', 0)

    elif i == 2:
        mean, max, min = crosscheck('b', 4)

    elif i == 3:
        mean, max, min = crosscheck('c', 1)

    elif i == 4:
        input('Please add aggressor b')
        mean, max, min = crosscheck('a', 0)

    elif i == 5:
        mean, max, min = crosscheck('c', 0)

    elif i == 6:
        mean, max, min = crosscheck('a', 4)

    elif i == 7:
        mean, max, min = crosscheck('c', 1)

    elif i == 8:
        input('Please add aggressor c')
        mean, max, min = crosscheck('a', 0)

    elif i == 9:
        mean, max, min = crosscheck('b', 0)

    elif i == 10:
        mean, max, min = crosscheck('a', 4)

    elif i == 11:
        mean, max, min = crosscheck('b', 4)

    ws1.cell(2 + i, 1, mean)
    ws1.cell(2 + i, 2, max)
    ws1.cell(2 + i, 3, min)

wb.save(path)
