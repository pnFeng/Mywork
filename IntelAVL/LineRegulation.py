import time
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from Mylib.bench_resource import brsc
from ICs import PineHurst
from bit_manipulate import bm
from openpyxl.styles import colors, Font


class Piht:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

    def enable(self, bEnable):
        if bEnable:
            self.dsh.i2c_write_reg(0x32, 0x80)
        else:
            self.dsh.i2c_write_reg(0x32, 0x00)

    def freq(self, ch, freq):
        if ch == 'a':
            ret, data = self.dsh.i2c_read_reg(0x29)
            if freq == 1000:
                data = bm.set_bit(data, 4)
                data = bm.clear_bit(data, 5)
                self.dsh.i2c_write_reg(0x29, data)
            elif freq == 750:
                data = bm.clear_bit(data, 5, 4)
                self.dsh.i2c_write_reg(0x29, data)
        elif ch == 'b':
            ret, data = self.dsh.i2c_read_reg(0x2A)
            if freq == 1000:
                data = bm.set_bit(data, 4)
                data = bm.clear_bit(data, 5)
                self.dsh.i2c_write_reg(0x2A, data)
            elif freq == 750:
                data = bm.clear_bit(data, 5, 4)
                self.dsh.i2c_write_reg(0x2A, data)
        elif ch == 'c':
            ret, data = self.dsh.i2c_read_reg(0x2A)
            if freq == 750:
                data = bm.clear_bit(data, 1, 0)
                self.dsh.i2c_write_reg(0x2A, data)
            elif freq == 1000:
                data = bm.set_bit(data, 0)
                data = bm.clear_bit(data, 1)
                self.dsh.i2c_write_reg(0x2A, data)
        print(self.dsh.i2c_read_reg(0x29), self.dsh.i2c_read_reg(0x2A))


path = './' + 'LineRegulation.xlsx'
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
m1 = brsc.DMM('USB0::0x2A8D::0x1401::MY57215694::0::INSTR')
m2 = brsc.DMM('USB0::0x2A8D::0x1401::MY57218437::0::INSTR')
p1.on(5, 2, 1)

wb = Workbook()
ws1 = wb.active

Rail = 'c'
channel = 3
a = Piht()


if Rail == 'a':
    a.dsh.i2c_write_reg(0x2f, 0x42)  # RAILA
    channel = 1
    load = [0.1, 1, 3, 4]
elif Rail == 'b':
    a.dsh.i2c_write_reg(0x2f, 0x12)  # RAILB
    channel = 2
    load = [0.1, 1, 3, 4]
elif Rail == 'c':
    a.dsh.i2c_write_reg(0x2f, 0x0A)  # RAILC
    channel = 3
    load = [0.1, 0.2, 0.5, 1]

a.freq(Rail, 750)
vin = [4.25, 4.5, 4.75, 5, 5.25, 5.5]
for j in vin:
    p1.on(j, 2, 1)
    for i in load:
        el1.cc(i, channel)
        el1.on([channel])
        vout = m2.meas()
        ws1.cell(row=load.index(i) + 1 + vin.index(j) * 5, column=1, value=j)
        ws1.cell(row=load.index(i) + 1 + vin.index(j) * 5, column=2, value=i)
        ws1.cell(row=load.index(i) + 1 + vin.index(j) * 5, column=3, value='750kHz')
        ws1.cell(row=load.index(i) + 1 + vin.index(j) * 5, column=4, value=vout)

a.freq(Rail, 1000)
vin = [4.25, 4.5, 4.75, 5, 5.25, 5.5]
for j in vin:
    p1.on(j, 2, 1)
    for i in load:
        el1.cc(i, channel)
        el1.on([channel])
        vout = m2.meas()
        ws1.cell(row=load.index(i) + 31 + vin.index(j) * 5, column=1, value=j)
        ws1.cell(row=load.index(i) + 31 + vin.index(j) * 5, column=2, value=i)
        ws1.cell(row=load.index(i) + 31 + vin.index(j) * 5, column=3, value='1000kHz')
        ws1.cell(row=load.index(i) + 31 + vin.index(j) * 5, column=4, value=vout)

wb.save(path)

el1.off([channel])
