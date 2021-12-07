# -*- coding: utf-8 -*-

import time

from bit_manipulate import bm
from openpyxl import Workbook
from Mylib.bench_resource import brsc
from ICs import PineHurst

p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
m2 = brsc.DMM('USB0::0x2A8D::0x1401::MY57218437::0::INSTR')


class SHVID:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(1000)

        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

    def enable(self):
        # enable
        ret, data = self.dsh.i2c_read_reg(0x2F)
        print('data', data, 'ret', ret)
        data = data | 0x04
        self.dsh.i2c_write_reg(0x2F, data)
        self.dsh.i2c_write_reg(0x32, 0x80)
        self.dsh.i2c_write_reg(0x6B, data)

    def VIDSetting(self, ch, VID, cre, freq):
        # self.dsh.i2c_write_reg(0x29, 0xC0)  # Forced SWA CCM
        # self.dsh.i2c_write_reg(0x2A, 0xCC)  # Forced SWB/C CCM
        # self.value = VID << 1 | cre
        if ch == 'a':
            self.dsh.i2c_write_reg(0x21, VID)
            ret, data = self.dsh.i2c_read_reg(0x29)
            if freq == 1000:
                data = bm.set_bit(data, 4)
                data = bm.clear_bit(data, 5)
            elif freq == 750:
                data = bm.clear_bit(data, 5, 4)
            self.dsh.i2c_write_reg(0x29, data)
        elif ch == 'b':
            self.dsh.i2c_write_reg(0x25, VID)
            ret, data = self.dsh.i2c_read_reg(0x2A)
            if freq == 1000:
                data = bm.set_bit(data, 4)
                data = bm.clear_bit(data, 5)
            elif freq == 750:
                data = bm.clear_bit(data, 5, 4)
            self.dsh.i2c_write_reg(0x2A, data)
        elif ch == 'c':
            self.dsh.i2c_write_reg(0x27, VID)
            ret, data = self.dsh.i2c_read_reg(0x2A)
            if freq == 750:
                data = bm.clear_bit(data, 1, 0)
            elif freq == 1000:
                data = bm.set_bit(data, 0)
                data = bm.clear_bit(data, 1)
            self.dsh.i2c_write_reg(0x2A, data)

    def VIDRB(self, rail):
        if rail == 'a':
            ret, data = a.dsh.i2c_read_reg(0x21)
        elif rail == 'b':
            ret, data = a.dsh.i2c_read_reg(0x25)
        elif rail == 'c':
            ret, data = a.dsh.i2c_read_reg(0x27)
        return ret, data


path = './' + 'VID.xlsx'
wb = Workbook()
ws1 = wb.active

ch = 'c'
channel = 3
freq = 1000

p1.on(5, 2, 1)
time.sleep(1)
a = SHVID()
a.enable()
time.sleep(1)

ws1.cell(1, 1, '#')
ws1.cell(1, 2, 'Multimeter VOUT(V)')
ws1.cell(1, 3, 'Setting')
ws1.cell(1, 4, 'Readback')

vol = [4.25, 5, 5.5]
for j in vol:
    p1.on(j, 2, 1)
    n = 128
    for i in range(n):
        print(i * 2, j)
        a.VIDSetting(ch, i * 2, 0, freq)
        time.sleep(0.5)
        ret, data = a.VIDRB(ch)
        meas1 = m2.meas()

        ws1.cell(i + 2 + 129 * vol.index(j), 1, str(i) + ' ' + str(j))
        ws1.cell(i + 2 + 129 * vol.index(j), 2, meas1)
        ws1.cell(i + 2 + 129 * vol.index(j), 3, hex(i * 2))
        ws1.cell(i + 2 + 129 * vol.index(j), 4, hex(data))

wb.save(path)
