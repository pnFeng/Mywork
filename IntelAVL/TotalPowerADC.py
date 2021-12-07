import time
from ICs import PineHurst
from openpyxl import Workbook
from bit_manipulate import bm
from Mylib.bench_resource import brsc


class PHCurMeas:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

        # 0x1A[1]=1
        ret, data = self.dsh.i2c_read_reg(0x1A)
        assert ret == 0
        data = bm.set_bit(data, 1)
        self.dsh.i2c_write_reg(0x1A, data)

        # 0x1b[6]=1
        ret, data = self.dsh.i2c_read_reg(0x1B)
        assert ret == 0
        data = bm.set_bit(data, 6)
        self.dsh.i2c_write_reg(0x1B, data)

    def enable(self):
        # enable
        ret, data = self.dsh.i2c_read_reg(0x2F)
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)
        self.dsh.i2c_write_reg(0x32, 0x80)

    def PowerMeas(self):
        ret, data = self.dsh.i2c_read_reg(0x0C)
        return data

    def VoltMeas(self):
        self.dsh.i2c_write_reg(0x30, 0x80)
        ret, volt_a = self.dsh.i2c_read_reg(0x31)
        self.dsh.i2c_write_reg(0x30, 0x90)
        ret, volt_b = self.dsh.i2c_read_reg(0x31)
        self.dsh.i2c_write_reg(0x30, 0x98)
        ret, volt_c = self.dsh.i2c_read_reg(0x31)
        return volt_a, volt_b, volt_c


path = './TotalPowerADC.xlsx'
step = 0.125
wb = Workbook()
ws1 = wb.active
ws1.cell(row=1, column=1, value='Load')
ws1.cell(row=1, column=2, value='Total Power')
ws1.cell(row=1, column=3, value='Total Power')
ws1.cell(row=1, column=4, value='Total Power')
ws1.cell(row=1, column=5, value='Total Power')
ws1.cell(row=1, column=6, value='Total Power')
ws1.cell(row=1, column=7, value='Total Power')
ws1.cell(row=1, column=8, value='Total Power')
ws1.cell(row=1, column=9, value='Total Power')
ws1.cell(row=1, column=10, value='Total Power')

el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
m2 = brsc.DMM('USB0::0x2A8D::0x1401::MY57218437::0::INSTR')
s1 = brsc.Scope('USB0::0x05FF::0x1023::3808N60392::INSTR')
p1.on(5, 4, 1)
p = PHCurMeas()
p.enable()

load = [0.5, 1, 2, 3, 4]
time.sleep(1)

for i in load:

    el1.cc(i, 1)
    el1.cc(i, 2)
    if i < 1.001:
        el1.cc(i, 3)
    else:
        el1.cc(0, 3)
    time.sleep(0.5)

    el1.on([1, 2, 3])

    time.sleep(0.5)

    vout = m2.meas()
    s1.scope.write('vbs app.measure.clearsweeps')
    time.sleep(1)
    for j in range(10):
        tp = p.PowerMeas()
        c1 = float(s1.scope.query('vbs? return = app.measure.p1.mean.result.value'))
        if j == 0:
            flag = 'nochange'
            last = tp
        elif last != tp:
            flag = 'change'
        r = load.index(i)
        ws1.cell(row=r + 2, column=1, value=i)
        ws1.cell(row=r + 2, column=2 + j, value=hex(tp)[2:])
        ws1.cell(row=r + 2, column=12 + j, value=c1)
        ws1.cell(row=r + 2, column=24, value=vout)
        print(i, j)
        time.sleep(1)

time.sleep(1)
el1.off([1, 2, 3])
wb.save(path)
