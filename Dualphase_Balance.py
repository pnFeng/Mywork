import time
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from Mylib.bench_resource import brsc
from ICs import PineHurst
from bit_manipulate import bm


class Piht:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

    def enable(self, bEnable):
        ret, data = self.dsh.i2c_read_reg(0x32)
        if bEnable:
            data = bm.set_bit(data, 7)
            self.dsh.i2c_write_reg(0x32, data)
        else:
            data = bm.clear_bit(data, 7)
            self.dsh.i2c_write_reg(0x32, data)


path = '../' + 'current_b.xlsx'
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
s1 = brsc.Scope('USB0::0x05FF::0x1023::3808N60392::INSTR')
# m1 = brsc.DMM('USB0::0x2A8D::0x1401::MY57215694::0::INSTR')
# m2 = brsc.DMM('USB0::0x2A8D::0x1401::MY57218437::0::INSTR')
p1.off(1)
p1.off(2)
time.sleep(1)
p1.on(3.3, 0.5, 1)
p1.on(12, 2, 2)

wb = Workbook()
ws1 = wb.active

a = Piht()
# a.dsh.i2c_write_reg(0x4f, 0x01)
# a.dsh.i2c_write_reg(0xac, 0x63)
# a.dsh.i2c_write_reg(0x8a, 0x02)
# a.dsh.i2c_write_reg(0x92, 0x87)
# a.dsh.i2c_write_reg(0x56, 0xcf)
# a.dsh.i2c_write_reg(0x2F, 0x42)
a.enable(True)
print(a.dsh.i2c_read_reg(0x32))

group = 0
j = 0

load = [0.0, 0.4, 0.7, 1.1, 1.4, 1.8, 2.1, 2.5, 2.8, 3.2, 3.5, 3.9, 4.2, 4.6, 4.9, 5.3, 5.6, 6.0, 6.3, 6.7, 7.0]

for i in load:
    j += 1
    print(i)
    el1.cc(i, 1)
    el1.on([1])
    time.sleep(0.5)
    s1.scope.write('vbs app.measure.clearsweeps')
    time.sleep(10)
    c1 = float(s1.scope.query('vbs? return = app.measure.p1.mean.result.value'))
    c2 = float(s1.scope.query('vbs? return = app.measure.p2.mean.result.value'))
    iout = el1.curmeas(1)
    c1_p = round(100 * c1 / (c1 + c2), 2)
    c2_p = round(100 * c2 / (c1 + c2), 2)
    # v2 = m1.meas()
    # v1 = m2.meas()

    ws1.cell(row=j + 1, column=1 + 8 * group, value=round(c1, 2))
    ws1.cell(row=j + 1, column=2 + 8 * group, value=c1_p)
    ws1.cell(row=j + 1, column=3 + 8 * group, value=round(c2, 2))
    ws1.cell(row=j + 1, column=4 + 8 * group, value=c2_p)
    ws1.cell(row=j + 1, column=5 + 8 * group, value=round(iout, 2))
    # ws1.cell(row=j + 1, column=6 + 8 * group, value=v1)
    # ws1.cell(row=j + 1, column=7 + 8 * group, value=v2)

    ws1.cell(row=1, column=1 + 8 * group, value='CHA')
    ws1.cell(row=1, column=2 + 8 * group, value='CHA_P')
    ws1.cell(row=1, column=3 + 8 * group, value='CHB')
    ws1.cell(row=1, column=4 + 8 * group, value='CHB_P')
    ws1.cell(row=1, column=5 + 8 * group, value='IOUT')
    # ws1.cell(row=1, column=6 + 8 * group, value='v1')
    # ws1.cell(row=1, column=7 + 8 * group, value='v2')

wb.save(path)

el1.off([1])
