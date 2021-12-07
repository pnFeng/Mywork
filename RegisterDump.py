import sys
import time
from openpyxl import Workbook
# import dolphin
from ICs import DolphinSH
from Mylib.bit_manipulate import bm
from Mylib.bench_resource import brsc
import prettytable as pt


class Piht:
    def __init__(self):
        self.dsh = DolphinSH()
        ret = self.dsh.connect()
        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        if ret is False:
            print("Failed to connect")
            sys.exit(-1)

        # self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

    def unlockDIMM(self, bEnable):
        if bEnable == 1:
            # # Un-lock DIMM vendor region(0x40~0x6f)
            self.dsh.i2c_write_reg(0x37, 0x73)
            self.dsh.i2c_write_reg(0x38, 0x94)
            self.dsh.i2c_write_reg(0x39, 0x40)

    def unlockIDT(self, bEnable):
        if bEnable == 1:
            # # Un-lock PMIC vendor region(0x70~0xff)
            self.dsh.i2c_write_reg(0x37, 0x79)
            self.dsh.i2c_write_reg(0x38, 0xbe)
            self.dsh.i2c_write_reg(0x39, 0x10)

    def securemode(self, bEnable):
        ret, data = self.dsh.i2c_read_reg(0x2F)
        print(ret, data)
        if bEnable == 0:
            data = bm.bit_manipulate(a, 0x2F, (2, 1))
        elif bEnable == 1:
            data = bm.bit_manipulate(0x2F, a, (2, 0))
        self.dsh.i2c_write_reg(0x2F, data)
        ret, data = self.dsh.i2c_read_reg(0x2F)
        print(ret, data)

    def enable(self, bEnable):
        if bEnable:
            self.dsh.i2c_write_reg(0x32, 0x80)
        else:
            self.dsh.i2c_write_reg(0x32, 0x00)


# path = './../RegisterDump1.xlsx'
path1 = './../RegisterDump1.txt'
# wb = Workbook()
# ws = wb.active

a = Piht()

a.securemode(0)
a.unlockDIMM(1)
a.unlockIDT(1)
a.enable(1)
time.sleep(1)

# ws.cell(row=1, column=1, value='Register')
# ws.cell(row=1, column=2, value='Value')
# for i in range(256):
#     ret, data = a.dsh.i2c_read_reg(i)
#     ws.cell(row=i + 2, column=1, value=hex(i))
#     ws.cell(row=i + 2, column=2, value=hex(data))
#
# wb.save(path)

f1 = open(path1, 'w')
tb = pt.PrettyTable()

headerlist = []
for i in range(16):
    headerlist.append(hex(i)[1:])
headerlist.insert(0, 'Reg')
tb.field_names = headerlist

for j in range(16):
    l = []
    for k in range(16):
        ret, data = a.dsh.i2c_read_reg(j * 16 + k)
        l.append(hex(data)[2:].zfill(2))
    l.insert(0, hex(j)[2:] + 'x')
    tb.add_row(l)
print(tb)
f1.write(str(tb))
