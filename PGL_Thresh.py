from ICs import PineHurst
import time
import sys
import visa
from bit_manipulate import bm
from openpyxl import Workbook, load_workbook
from bit_manipulate import bm
from Mylib.bench_resource import brsc


class SHVID:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        if ret is False:
            print("Failed to connect")
            sys.exit(-1)

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

    def enable(self):
        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        # Diable OVP
        self.dsh.i2c_write_reg(0x56, 0xF0)

        # enable
        self.dsh.i2c_write_reg(0x2F, 0x06)
        self.dsh.i2c_write_reg(0x32, 0x80)

    def Trim_REF(self, binary):
        ret0, data = self.dsh.i2c_read_reg(rgt)
        data_2bit = bm.get_bit(data, 'l', [7, 6])
        data = data_2bit[0] * 2 ** 7 + data_2bit[1] * 2 ** 6 + binary
        self.dsh.i2c_write_reg(rgt, data)
        return data

    def PGL_UV(self, railx, sel, modex):
        if modex == 'pgl':
            # unMask PG
            ret, data = self.dsh.i2c_read_reg(0x15)
            data = bm.clear_bit(data, 5, 3, 2)
            self.dsh.i2c_write_reg(0x15, 0x80)

            if railx == 'a':
                rgt = 0x21
            elif railx == 'b':
                rgt = 0x25
            elif railx == 'c':
                rgt = 0x27

            if sel == 0:
                ret, data = self.dsh.i2c_read_reg(rgt)
                data = bm.clear_bit(data, 0)
                self.dsh.i2c_write_reg(rgt, data)
            elif sel == 1:
                ret, data = self.dsh.i2c_read_reg(rgt)
                data = bm.set_bit(data, 0)
                self.dsh.i2c_write_reg(rgt, data)

            ret, data = self.dsh.i2c_read_reg(rgt)
            print
            'pgl register {}'.format(bin(data))

        elif modex == 'uv':
            # Mask PG
            ret, data = self.dsh.i2c_read_reg(0x15)
            data = bm.set_bit(data, 5, 3, 2)
            self.dsh.i2c_write_reg(0x15, 0x80)

            if railx == 'a':
                rgt = 0x22
            elif railx == 'b':
                rgt = 0x26
            elif railx == 'c':
                rgt = 0x28

            if sel == 0:
                ret, data = self.dsh.i2c_read_reg(rgt)
                data = bm.clear_bit(data, 3, 2)
                self.dsh.i2c_write_reg(rgt, data)
            elif sel == 1:
                ret, data = self.dsh.i2c_read_reg(rgt)
                data = bm.clear_bit(data, 3)
                data = bm.set_bit(data, 2)
                self.dsh.i2c_write_reg(rgt, data)
            elif sel == 2:
                ret, data = self.dsh.i2c_read_reg(rgt)
                data = bm.set_bit(data, 3)
                data = bm.clear_bit(data, 2)
                self.dsh.i2c_write_reg(rgt, data)
            elif sel == 3:
                ret, data = self.dsh.i2c_read_reg(rgt)
                data = bm.set_bit(data, 3, 2)
                self.dsh.i2c_write_reg(rgt, data)

            ret, data = self.dsh.i2c_read_reg(rgt)
            print
            'uv register {}'.format(bin(data))

    def opsetting(self, ch, vol):
        if ch == 'a':
            data = 2 * round((vol - 0.8) / 0.005, 0)
            self.dsh.i2c_write_reg(0x21, data)
        elif ch == 'b':
            data = 2 * round((vol - 0.8) / 0.005, 0)
            self.dsh.i2c_write_reg(0x25, data)
        elif ch == 'c':
            data = 2 * round((vol - 1.5) / 0.005, 0)
            self.dsh.i2c_write_reg(0x27, data)


rail = 'c'

if rail == 'a':
    rgt = 0x93
    list1 = [0.8, 1.1, 1.435]
elif rail == 'b':
    rgt = 0xa1
    list1 = [0.8, 1.1, 1.435]
elif rail == 'c':
    rgt = 0xa9
    list1 = [1.5, 1.8, 2.135]

mode = 'pgl'  # pgl
path = './' + mode + '_' + rail + '.xlsx'
m1 = brsc.DMM('USB0::0x2A8D::0x1401::MY57215694::0::INSTR')
m2 = brsc.DMM('USB0::0x2A8D::0x1401::MY57218437::0::INSTR')
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
a = SHVID()

wb = Workbook()
ws1 = wb.active

list2 = [0, 1]  # i

for j in list1:
    cnt1 = list1.index(j)
    ws1.cell(row=1, column=1 + 6 * cnt1, value=j)
    ws1.cell(row=1, column=2 + 6 * cnt1, value='PG_L')
    for i in list2:
        p1.off()
        time.sleep(0.5)
        p1.output()
        time.sleep(0.5)
        a.opsetting(rail, j)
        a.PGL_UV(rail, i, mode)
        a.enable()
        ret, data = a.dsh.i2c_read_reg(rgt)
        data = bm.get_bit(data, 's', range(5, -1, -1))
        data = int(data, 2)
        init_value = bm.binary2value(bm.orgncode2cpmcode(data, 5), 5)
        for k in range(init_value + 2 ** 5 + 1):
            setting = init_value
            init_value -= 1
            setting = bm.orgncode2cpmcode(bm.value2binary(setting, 5), 5)
            val = a.Trim_REF(setting)
            meas2 = m2.meas()
            meas1 = m1.meas()
            print
            'j {} i {} k {} val {} meas1 {} meas2 {}'.format(j, i, k, val, meas1, meas2)
            if meas1 < 0.2:
                print('threshold found')
                # record the data in excel
                ws1.cell(row=2 + i, column=2 + 6 * cnt1, value=meas2)
                ret, data = a.dsh.i2c_read_reg(0x08)
                data = bm.get_bit(data, 's', [5, 3, 2])
                ws1.cell(row=2 + i, column=3 + 6 * cnt1, value=data)
                wb.save(path)
                break
