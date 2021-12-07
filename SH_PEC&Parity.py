import time
from openpyxl import Workbook
from Mylib.bench_resource import brsc
from ICs import PineHurst
from bit_manipulate import bm
import sys
from dolphin import *

path = './../' + 'SH_Event_Parity&PEC.xlsx'
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
m1 = brsc.DMM('USB0::0x2A8D::0x1401::MY57215694::0::INSTR')
m2 = brsc.DMM('USB0::0x2A8D::0x1401::MY57218437::0::INSTR')
status_rgt = [0x08, 0x09, 0x0a, 0x0b, 0x33]
clear_rgt = [0x10, 0x11, 0x12, 0x13, 0x14]
mask_rgt = [0x15, 0x16, 0x17, 0x18, 0x19]
name = 'None'
statusmap = ['R8[7]VIN_PG', 'R8[6]CTS', 'R8[5]SWA_PG', 'R8[4]SWB_PG', 'R8[3]SWC_PG', 'R8[2]SWD_PG', 'R8[1]VM_OV',
             'R8[0]VIN_OV',
             'R9[7]VIN_HTW', 'R9[6]VB_PG', 'R9[5]LDO1.8PG', 'R9[4]VM2VIN_SO', 'R9[3]SWAHiCurWar', 'R9[2]SWBHiCurWar',
             'R9[1]SWCHiCurWar', 'R9[0]SWDHiCurWar',
             'RA[7]SWAOV', 'RA[6]SWBOV', 'RA[5]SWCOV', 'RA[4]SWDOV', 'RA[3]PECErr', 'RA[2]ParityErr',
             'RA[1]GlobalStatus', 'RA[0]RSV',
             'RB[7]SWACurLim', 'RB[6]SWBCurLim', 'RB[5]SWCCurLim', 'RB[4]SWDCurLim', 'RB[3]SWAUV', 'RB[2]SWBUV',
             'RB[1]SWCUV', 'RB[0]SWDUV',
             'R33[7]TM', 'R33[6]TM', 'R33[5]TM', 'R33[4]VM_PG_SOMO', 'R33[3]VINUV', 'R33[2]LDO1.1PG', 'R33[1]RSV',
             'R33[0]RSV']
maskmap = ['R15[7]VIN_PG', 'R15[6]RSV', 'R15[5]SWAPG', 'R15[4]SWBPG', 'R15[3]SWCPG', 'R15[2]SWDPG', 'R15[1]VM_OV',
           'R15[0]VINOV',
           'R16[7]HTW', 'R16[6]VB_PG', 'R16[5]LDO1.8PG', 'R16[4]VM2VIN_SO', 'R16[3]SWAHICURWAN', 'R16[2]SWBHICURWAN',
           'R16[1]SWCHICURWAN', 'R16[0]SWDHICURWAN',
           'R17[7]SWAOV', 'R17[6]SWBOV', 'R17[5]SWCOV', 'R17[4]SWDOV', 'R17[3]PECERR', 'R17[2]PARITYERR', 'R17[1]RSV',
           'R17[0]RSV',
           'R18[7]SWACURLIM', 'R18[6]SWBCURLIM', 'R18[5]SWCCURLIM', 'R18[4]SWDCURLIM', 'R18[3]SWAUV', 'R18[2]SWBUV',
           'R18[1]SWCUV', 'R18[0]SWDUV',
           'R19[7]RSV', 'R19[6]RSV', 'R19[5]RSV', 'R19[4]VM_PG_SOM', 'R19[3]VB&VIN_UV', 'R19[2]LDO1.0PG', 'R19[1]RSV',
           'R19[0]RSV']


class Piht:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

    def unlock(self):
        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

    def GSI(self, bEnable):
        ret, data = self.dsh.i2c_read_reg(0x1B)
        assert ret == 0
        if bEnable == True:
            data = bm.set_bit(data, 3)
        elif bEnable == False:
            data = bm.clear_bit(data, 3)
        self.dsh.i2c_write_reg(0x1B, data)

    def globalclear(self):
        self.dsh.i2c_write_reg(0x14, 0x01)

    def enable(self, bEnable):
        if bEnable:
            self.dsh.i2c_write_reg(0x32, 0x80)
        else:
            self.dsh.i2c_write_reg(0x32, 0x00)

    def MaskCtrl(self, Maskctrl):
        ret, data = self.dsh.i2c_read_reg(0x2F)
        if Maskctrl == 0b00:
            data = bm.clear_bit(data, 0, 1)
        elif Maskctrl == 0b01:
            data = bm.clear_bit(data, 1)
            data = bm.set_bit(data, 0)
        elif Maskctrl == 0b10:
            data = bm.set_bit(data, 1)
            data = bm.clear_bit(data, 0)
        elif Maskctrl == 0b11:
            print('Please Check MastCtrl')
            raise
        self.dsh.i2c_write_reg(0x2F, data)


def rgtset(dongle, rgt, set, *n):
    ret, data = dongle.dsh.i3c_read_reg(rgt)
    if set == 1:
        data = bm.set_bit(data, *n)
    elif set == 0:
        data = bm.clear_bit(data, *n)
    dongle.dsh.i3c_write_reg(rgt, data)


def statuscheck(dongle):
    rgt = []
    statusindex = ''
    for i in status_rgt:
        ret, data = dongle.dsh.i3c_read_reg(i)
        if ret != 0:
            rgt.append(data)
        elif ret == 0:
            rgt.append(hex(data))
            index = status_rgt.index(i)
            str1 = str(bin(data))[2:10].zfill(8)
            for index1, s in enumerate(str1):
                if s == '1':
                    statusindex = statusindex + ';' + str(statusmap[index1 + index * 8])

        VOUTA = el1.volmeas(1)
        VOUTB = el1.volmeas(2)
        VOUTC = el1.volmeas(3)
        VOUTD = el1.volmeas(4)
        Iin = p1.curmeas(2)
        PG = m1.meas()
        GSI = m2.meas()

        if VOUTA < 0.3 or VOUTB < 0.3 or VOUTC < 0.3 or VOUTD < 0.3 or Iin < 0.001:
            VR = 'disabled'
        else:
            VR = 'enabled'

        if PG > 1.6:
            PG = 'H'
        elif PG < 0.3:
            PG = 'L'
        else:
            PG = 'Please Check'

        if GSI > 1.6:
            GSI = 'H'
        elif GSI < 0.3:
            GSI = 'L'
        else:
            GSI = 'Please Check'

    return rgt, VR, PG, GSI, statusindex


def PEC(dongle, bMask):
    if bMask == 0:
        rgtset(dongle, 0x17, 0, 3)
    elif bMask == 1:
        rgtset(dongle, 0x17, 1, 3)
    dongle.dsh.ccc_SETAASA()
    dongle.dsh.util_i3c_modify_reg_bits(0x34, [(7, 1)])
    ret, dat = dongle.dsh.i3c_read_reg(0x34)
    dongle.dsh.bEnPEC = True
    dongle.dsh.i3c_write_reg(0x32, 0x80)
    dongle.dsh.inject_error(DP_INJECT_PEC_ERROR)
    ret = dongle.dsh.i3c_write_reg(0x2F, 0x80)
    assert ret == 0
    dongle.dsh.inject_error(DP_INJECT_NONE_ERROR)


def clear_PEC(dongle):
    rgt = []
    statusindex = ''
    flag = 0
    rgtset(dongle, 0x12, 1, 3)  # clear PEC
    for i in status_rgt:
        ret, data = dongle.dsh.i3c_read_reg(i)
        flag = flag + data
        rgt.append(hex(data))
        index = status_rgt.index(i)
        str1 = str(bin(data))[2:10].zfill(8)
        for index1, s in enumerate(str1):
            if s == '1':
                statusindex = statusindex + ';' + str(statusmap[index1 + index * 8])
    print('Clear flag:{}; {}; {}'.format(flag, rgt, statusindex))
    return flag, statusindex


def Parity(dongle, bMask):
    if bMask == 0:
        rgtset(dongle, 0x17, 0, 2)
    elif bMask == 1:
        rgtset(dongle, 0x17, 1, 2)
    dongle.dsh.ccc_SETAASA()
    dongle.dsh.i3c_write_reg(0x32, 0x80)
    dongle.dsh.inject_error(DP_INJECT_PARITY_ERROR)
    ret = dongle.dsh.i3c_write_reg(0x2F, 0x80)
    assert ret == 0
    dongle.dsh.inject_error(DP_INJECT_NONE_ERROR)


def clear_Parity(dongle):
    rgt = []
    statusindex = ''
    flag = 0
    rgtset(dongle, 0x12, 1, 2)  # clear PEC
    for i in status_rgt:
        ret, data = dongle.dsh.i3c_read_reg(i)
        flag = flag + data
        rgt.append(hex(data))
        index = status_rgt.index(i)
        str1 = str(bin(data))[2:10].zfill(8)
        for index1, s in enumerate(str1):
            if s == '1':
                statusindex = statusindex + ';' + str(statusmap[index1 + index * 8])
    print('Clear flag:{}; {}; {}'.format(flag, rgt, statusindex))
    return flag, statusindex


def main():
    wb = Workbook()
    ws1 = wb.active

    ws1.cell(row=1, column=1, value='Item')
    ws1.cell(row=1, column=2, value='Status Register')
    ws1.cell(row=1, column=3, value='VR')
    ws1.cell(row=1, column=4, value='PG')
    ws1.cell(row=1, column=5, value='GSI')
    ws1.cell(row=1, column=6, value='Status Value')
    ws1.cell(row=1, column=7, value='Clear Flag')
    ws1.cell(row=1, column=8, value='VR_Cleared')
    ws1.cell(row=1, column=9, value='PG_Cleared')
    ws1.cell(row=1, column=10, value='GSI_Cleared')

    a = Piht()
    mask = 1

    p1.off(1)
    p1.off(2)
    time.sleep(1)
    p1.on(3.3, 1, 1)
    p1.on(12, 2, 2)
    time.sleep(1)

    a.unlock()
    a.GSI(True)
    a.MaskCtrl(0b10)

    # PEC(a, mask)
    Parity(a, mask)
    status, VR, PG, GSI, statusvalue = statuscheck(a)
    print('{} status: {}; VR: {}; PG: {}; GSI: {};statusvalue: {}'.format(name, status, VR, PG, GSI, statusvalue))
    # flag, statusindex = clear_PEC(a)
    flag, statusindex = clear_Parity(a)
    statusC, VRC, PGC, GSIC, statusvalueC = statuscheck(a)
    print('{} statusC: {}; VRC: {}; PGC: {}; GSIC: {};statusvalueC: {}'.format(name, statusC, VRC, PGC, GSIC,
                                                                               statusvalueC))

    ws1.cell(row=2, column=1, value=name)
    ws1.cell(row=2, column=2, value=(str(status[0]) + ',' + str(status[1])) + ',' + str(status[2]) + ',' + str(
        status[3]) + ',' + str(status[4]))
    ws1.cell(row=2, column=3, value=VR)
    ws1.cell(row=2, column=4, value=PG)
    ws1.cell(row=2, column=5, value=GSI)
    ws1.cell(row=2, column=6, value=statusvalue)
    ws1.cell(row=2, column=7, value=flag)
    ws1.cell(row=2, column=8, value=VRC)
    ws1.cell(row=2, column=9, value=PGC)
    ws1.cell(row=2, column=10, value=GSIC)

    wb.save(path)


if __name__ == "__main__":
    main()
