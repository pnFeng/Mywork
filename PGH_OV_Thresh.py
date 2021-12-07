import time
import visa
from bit_manipulate import bm
from openpyxl import Workbook
from ICs import PineHurst
from Mylib.bench_resource import brsc


class Piht:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

    def enable(self):
        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        # disable ooa and
        ret, data = self.dsh.i2c_read_reg(0x82)
        data = bm.clear_bit(data, 4)
        self.dsh.i2c_write_reg(0x82, data)

        # enable
        ret, data = self.dsh.i2c_read_reg(0x2F)
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)
        self.dsh.i2c_write_reg(0x32, 0x80)

    def PGH_OV_Th(self, railx, sel, modex):
        if railx == 'a':
            rgt = 0x22
        elif railx == 'b':
            rgt = 0x26
        elif railx == 'c':
            rgt = 0x28

        # disable soft OV
        ret, data = self.dsh.i2c_read_reg(0x82)
        data = bm.clear_bit(data, 1)
        self.dsh.i2c_write_reg(0x82, data)

        if modex == 'ov':
            # Mask PG_H
            ret, data = self.dsh.i2c_read_reg(0x15)
            data = bm.set_bit(data, 5, 3, 2)
            self.dsh.i2c_write_reg(0x15, data)

            if sel == 0:
                ret, data = self.dsh.i2c_read_reg(rgt)
                data = bm.clear_bit(data, 5, 4)
                self.dsh.i2c_write_reg(rgt, data)
            elif sel == 1:
                ret, data = self.dsh.i2c_read_reg(rgt)
                data = bm.clear_bit(data, 5)
                data = bm.set_bit(data, 4)
                self.dsh.i2c_write_reg(rgt, data)
            elif sel == 2:
                ret, data = self.dsh.i2c_read_reg(rgt)
                data = bm.set_bit(data, 5)
                data = bm.clear_bit(data, 4)
                self.dsh.i2c_write_reg(rgt, data)

        elif modex == 'pg_h':
            # unMask PG_H
            ret, data = self.dsh.i2c_read_reg(0x15)
            data = bm.clear_bit(data, 5, 3, 2)
            self.dsh.i2c_write_reg(0x15, data)

            if sel == 0:
                ret, data = self.dsh.i2c_read_reg(rgt)
                data = bm.clear_bit(data, 7, 6)
                self.dsh.i2c_write_reg(rgt, data)
            elif sel == 1:
                ret, data = self.dsh.i2c_read_reg(rgt)
                data = bm.clear_bit(data, 7)
                data = bm.set_bit(data, 6)
                self.dsh.i2c_write_reg(rgt, data)
            elif sel == 2:
                ret, data = self.dsh.i2c_read_reg(rgt)
                data = bm.set_bit(data, 7)
                data = bm.clear_bit(data, 6)
                self.dsh.i2c_write_reg(rgt, data)

    def opsetting(self, ch, vol):
        if ch == 'a':
            data = 2 * round((vol - 0.8) / 0.005, 0)
            self.dsh.i2c_write_reg(0x21, int(data))
        elif ch == 'b':
            data = 2 * round((vol - 0.8) / 0.005, 0)
            self.dsh.i2c_write_reg(0x25, int(data))
        elif ch == 'c':
            data = 2 * round((vol - 1.5) / 0.005, 0)
            self.dsh.i2c_write_reg(0x27, int(data))


rail = 'c'
mode = 'ov'  # pg_h ov
path = '../' + mode + '_' + rail + '.xlsx'
m1 = brsc.DMM('USB0::0x2A8D::0x1401::MY57215694::0::INSTR')
m2 = brsc.DMM('USB0::0x2A8D::0x1401::MY57218437::0::INSTR')
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
a = Piht()

wb = Workbook()
ws1 = wb.active

coarse_step = 0.025
fine_step = 0.001

list1 = [1.5, 1.8, 2.135]  # j  [0.8, 1.1, 1.435] [1.5, 1.8, 2.135]
list2 = [0, 1, 2]  # i

p1.off()
time.sleep(1)

for j in list1:
    cnt1 = list1.index(j)
    ws1.cell(row=1, column=1 + 6 * cnt1, value=j)
    ws1.cell(row=1, column=3 + 6 * cnt1, value='PG Status')
    ws1.cell(row=1, column=4 + 6 * cnt1, value='OV status')
    for i in list2:
        coarseflag = True
        setting = 0
        hitflag = 0
        abnormalflag = 0
        while 1:
            if (hitflag == 1) | (abnormalflag == 1):
                break
            else:
                # initial
                p1.on(5, 2, 2)
                p1.on(j, 1, 1)  # start with 0.8 target

                time.sleep(1)
                a.opsetting(rail, j)  # output setting should before PGTH setting
                a.PGH_OV_Th(rail, i, mode)
                a.enable()
                time.sleep(1)
                setting_fine = setting - coarse_step
                meas1 = m1.meas()
                meas2 = m2.meas()  # PG signal
                setting = meas1
                if meas2 < 1.2:
                    print
                    'PMIC Abnormal'
                    abnormalflag = 1
                hi = 1.2 * j
                lo = 0.6 * j
                cnt2 = list2.index(i)

            # find the target
            while 1:
                a.PGH_OV_Th(rail, i, mode)
                if coarseflag == True:  # coarse tuning
                    meas1 = m1.meas()
                    p1.on(setting, 1, 1)
                    print('Coarse Addressing', 'setting', setting, 'i', i, 'j', j)
                    meas2 = m2.meas()
                    if meas2 < 0.2:
                        print('coarse range found')
                        coarseflag = False
                        p1.off()
                        time.sleep(1)
                        break
                    else:
                        setting = setting + coarse_step
                        if meas1 >= hi or meas1 <= lo:
                            print('thresold not found')
                            raise NameError('Exceed high limit')

                elif coarseflag == False:  # fine tuning
                    meas1 = m1.meas()
                    p1.on(setting_fine, 1, 1)
                    print('Fine Addressing', 'setting', setting_fine, 'i', i, 'j', j)
                    meas2 = m2.meas()
                    if meas2 < 0.2:
                        print('threshold found')
                        # record the data in excel
                        ws1.cell(row=2 + cnt2, column=2 + 6 * cnt1, value=meas1)
                        ret, data = a.dsh.i2c_read_reg(0x08)
                        data = bm.get_bit(data, 's', [5, 3, 2])
                        ws1.cell(row=2 + cnt2, column=3 + 6 * cnt1, value=data)

                        ret, data = a.dsh.i2c_read_reg(0x0a)
                        data = bm.get_bit(data, 's', [7, 5, 4])
                        ws1.cell(row=2 + cnt2, column=4 + 6 * cnt1, value=data)

                        p1.off()
                        time.sleep(1)
                        wb.save(path)
                        hitflag = 1
                        break
                    else:
                        setting_fine += fine_step
                        if meas1 >= hi or meas1 <= lo:
                            print('thresold not found')
                            raise NameError('Exceed limit')
