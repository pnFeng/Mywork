import time
from openpyxl import Workbook
from Mylib.bench_resource import brsc
from ICs import PineHurst
from bit_manipulate import bm
import sys
from dolphin import *

path = './../' + 'SH_ErrorInjection.xlsx'
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
m1 = brsc.DMM('USB0::0x2A8D::0x1401::MY57215694::0::INSTR')
m2 = brsc.DMM('USB0::0x2A8D::0x1401::MY57218437::0::INSTR')
status_rgt = [0x08, 0x09, 0x0a, 0x0b, 0x33]
clear_rgt = [0x10, 0x11, 0x12, 0x13, 0x14]
mask_rgt = [0x15, 0x16, 0x17, 0x18, 0x19]
name = 'None'
statusmap = ['R8[7]VIN_PG', 'R8[6]CTS', 'R8[5]SWA_PG', 'R8[4]SWB_PG', 'R8[3]SWC_PG', 'R8[2]SWD_PG', 'R8[1]VM_OV',
             'R8[0]VIN_OV',
             'R9[7]VIN_HTW', 'R9[6]VB_PG', 'R9[5]LDO1.8PG', 'R9[4]VM2VIN_SO', 'R9[3]SWAHiCurWar', 'R9[2]SWBHiCurWar',
             'R9[1]SWCHiCurWar', 'R9[0]SWDHiCurWar',
             'RA[7]SWAOV', 'RA[6]SWBOV', 'RA[5]SWCOV', 'RA[4]SWDOV', 'RA[3]PECErr', 'RA[2]ParityErr',
             'RA[1]GlobalStatus', 'RA[0]RSV',
             'RB[7]SWACurLim', 'RB[6]SWBCurLim', 'RB[5]SWCCurLim', 'RB[4]SWDCurLim', 'RB[3]SWAUV', 'RB[2]SWBUV',
             'RB[1]SWCUV', 'RB[0]SWDUV',
             'R33[7]TM', 'R33[6]TM', 'R33[5]TM', 'R33[4]VM_PG_SOMO', 'R33[3]VINUV', 'R33[2]LDO1.1PG', 'R33[1]RSV',
             'R33[0]RSV']
maskmap = ['R15[7]VIN_PG', 'R15[6]RSV', 'R15[5]SWAPG', 'R15[4]SWBPG', 'R15[3]SWCPG', 'R15[2]SWDPG', 'R15[1]VM_OV',
           'R15[0]VINOV',
           'R16[7]HTW', 'R16[6]VB_PG', 'R16[5]LDO1.8PG', 'R16[4]VM2VIN_SO', 'R16[3]SWAHICURWAN', 'R16[2]SWBHICURWAN',
           'R16[1]SWCHICURWAN', 'R16[0]SWDHICURWAN',
           'R17[7]SWAOV', 'R17[6]SWBOV', 'R17[5]SWCOV', 'R17[4]SWDOV', 'R17[3]PECERR', 'R17[2]PARITYERR', 'R17[1]RSV',
           'R17[0]RSV',
           'R18[7]SWACURLIM', 'R18[6]SWBCURLIM', 'R18[5]SWCCURLIM', 'R18[4]SWDCURLIM', 'R18[3]SWAUV', 'R18[2]SWBUV',
           'R18[1]SWCUV', 'R18[0]SWDUV',
           'R19[7]RSV', 'R19[6]RSV', 'R19[5]RSV', 'R19[4]VM_PG_SOM', 'R19[3]VB&VIN_UV', 'R19[2]LDO1.0PG', 'R19[1]RSV',
           'R19[0]RSV']


class Piht:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

    def unlock(self):
        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

    def GSI(self, bEnable):
        ret, data = self.dsh.i2c_read_reg(0x1B)
        assert ret == 0
        if bEnable == True:
            data = bm.set_bit(data, 3)
        elif bEnable == False:
            data = bm.clear_bit(data, 3)
        self.dsh.i2c_write_reg(0x1B, data)

    def globalclear(self):
        self.dsh.i2c_write_reg(0x14, 0x01)

    def enable(self, bEnable):
        if bEnable:
            self.dsh.i2c_write_reg(0x32, 0x80)
        else:
            self.dsh.i2c_write_reg(0x32, 0x00)

    def MaskCtrl(self, Maskctrl):
        ret, data = self.dsh.i2c_read_reg(0x2F)
        if Maskctrl == 0b00:
            data = bm.clear_bit(data, 0, 1)
        elif Maskctrl == 0b01:
            data = bm.clear_bit(data, 1)
            data = bm.set_bit(data, 0)
        elif Maskctrl == 0b10:
            data = bm.set_bit(data, 1)
            data = bm.clear_bit(data, 0)
        elif Maskctrl == 0b11:
            print('Please Check MastCtrl')
            raise
        self.dsh.i2c_write_reg(0x2F, data)


class Ej:

    def SWA_OV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x17, 0, 7)
        elif bMask == 1:
            rgtset(dongle, 0x17, 1, 7)
        dongle.dsh.i2c_write_reg(0x35, 0x90)
        global name
        name = sys._getframe().f_code.co_name

    def SWB_OV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x17, 0, 6)
        elif bMask == 1:
            rgtset(dongle, 0x17, 1, 6)
        dongle.dsh.i2c_write_reg(0x35, 0xA0)
        global name
        name = sys._getframe().f_code.co_name

    def SWC_OV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x17, 0, 5)
        elif bMask == 1:
            rgtset(dongle, 0x17, 1, 5)
        dongle.dsh.i2c_write_reg(0x35, 0xB0)
        global name
        name = sys._getframe().f_code.co_name

    def SWD_OV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x17, 0, 4)
        elif bMask == 1:
            rgtset(dongle, 0x17, 1, 4)
        dongle.dsh.i2c_write_reg(0x35, 0xC0)
        global name
        name = sys._getframe().f_code.co_name

    def SWAUV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x18, 0, 3)
        elif bMask == 1:
            rgtset(dongle, 0x18, 1, 3)
        dongle.dsh.i2c_write_reg(0x35, 0x98)
        global name
        name = sys._getframe().f_code.co_name

    def SWBUV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x18, 0, 2)
        elif bMask == 1:
            rgtset(dongle, 0x18, 1, 2)
        dongle.dsh.i2c_write_reg(0x35, 0xA8)
        global name
        name = sys._getframe().f_code.co_name

    def SWCUV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x18, 0, 1)
        elif bMask == 1:
            rgtset(dongle, 0x18, 1, 1)
        dongle.dsh.i2c_write_reg(0x35, 0xB8)
        global name
        name = sys._getframe().f_code.co_name

    def SWDUV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x18, 0, 0)
        elif bMask == 1:
            rgtset(dongle, 0x18, 1, 0)
        dongle.dsh.i2c_write_reg(0x35, 0xC8)
        global name
        name = sys._getframe().f_code.co_name

    def VINOV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x15, 0, 0)
        elif bMask == 1:
            rgtset(dongle, 0x15, 1, 0)
        dongle.dsh.i2c_write_reg(0x35, 0xD0)
        global name
        name = sys._getframe().f_code.co_name

    def VINUV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x19, 0, 3)  # mask VIN/Vbias UV
            rgtset(dongle, 0x16, 0, 6)  # mask Vbias PG
        elif bMask == 1:
            rgtset(dongle, 0x19, 1, 3)
            rgtset(dongle, 0x16, 1, 6)
        dongle.dsh.i2c_write_reg(0x35, 0xD8)
        global name
        name = sys._getframe().f_code.co_name

    def VMOV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x15, 0, 1)
        elif bMask == 1:
            rgtset(dongle, 0x15, 1, 1)
        dongle.dsh.i2c_write_reg(0x35, 0xE0)
        global name
        name = sys._getframe().f_code.co_name

    def All_OV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x15, 0, 1)
        elif bMask == 1:
            rgtset(dongle, 0x15, 1, 1)
        dongle.dsh.i2c_write_reg(0x35, 0xF0)
        global name
        name = sys._getframe().f_code.co_name

    def All_UV(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x15, 0, 1)
        elif bMask == 1:
            rgtset(dongle, 0x15, 1, 1)
        dongle.dsh.i2c_write_reg(0x35, 0xF8)
        global name
        name = sys._getframe().f_code.co_name

    def VM_VIN_SOM(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x16, 0, 4)
        elif bMask == 1:
            rgtset(dongle, 0x16, 1, 4)
        dongle.dsh.i2c_write_reg(0x35, 0x81)
        global name
        name = sys._getframe().f_code.co_name

    def CTS(self, dongle, bMask):
        dongle.dsh.i2c_write_reg(0x35, 0x82)
        global name
        name = sys._getframe().f_code.co_name

    def HTW(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x16, 0, 7)
        elif bMask == 1:
            rgtset(dongle, 0x16, 1, 7)
        dongle.dsh.i2c_write_reg(0x35, 0x83)
        global name
        name = sys._getframe().f_code.co_name

    def LDO1p8PG(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x16, 0, 5)
        elif bMask == 1:
            rgtset(dongle, 0x16, 1, 5)
        dongle.dsh.i2c_write_reg(0x35, 0x84)
        global name
        name = sys._getframe().f_code.co_name

    def HCW(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x16, 0, 3)
            rgtset(dongle, 0x16, 0, 2)
            rgtset(dongle, 0x16, 0, 1)
            rgtset(dongle, 0x16, 0, 0)
        elif bMask == 1:
            rgtset(dongle, 0x16, 1, 3)
            rgtset(dongle, 0x16, 1, 2)
            rgtset(dongle, 0x16, 1, 1)
            rgtset(dongle, 0x16, 1, 0)
        dongle.dsh.i2c_write_reg(0x35, 0x85)
        global name
        name = sys._getframe().f_code.co_name

    def CAMP(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x16, 0, 5)
        elif bMask == 1:
            rgtset(dongle, 0x16, 1, 5)
        rgtset(dongle, 0x32, 1, 5)  # CAMP(PG) IO PIN
        dongle.dsh.i2c_write_reg(0x35, 0x86)
        global name
        name = sys._getframe().f_code.co_name

    def CurLim(self, dongle, bMask):
        if bMask == 0:
            rgtset(dongle, 0x18, 0, 7)
            rgtset(dongle, 0x18, 0, 6)
            rgtset(dongle, 0x18, 0, 5)
            rgtset(dongle, 0x18, 0, 4)
        elif bMask == 1:
            rgtset(dongle, 0x18, 1, 7)
            rgtset(dongle, 0x18, 1, 6)
            rgtset(dongle, 0x18, 1, 5)
            rgtset(dongle, 0x18, 1, 4)
        dongle.dsh.i2c_write_reg(0x35, 0x87)
        global name
        name = sys._getframe().f_code.co_name


class clear_event:
    def clear_SWA_OV(self, dongle):
        rgtset(dongle, 0x12, 1, 7)  # clear SWA OV

    def clear_SWB_OV(self, dongle):
        rgtset(dongle, 0x12, 1, 6)  # clear SWB OV

    def clear_SWC_OV(self, dongle):
        rgtset(dongle, 0x12, 1, 5)  # clear SWC OV

    def clear_SWD_OV(self, dongle):
        rgtset(dongle, 0x12, 1, 4)  # clear SWD OV

    def clear_SWA_UV(self, dongle):
        rgtset(dongle, 0x13, 1, 3)  # clear SWA UV

    def clear_SWB_UV(self, dongle):
        rgtset(dongle, 0x13, 1, 2)  # clear SWB UV

    def clear_SWC_UV(self, dongle):
        rgtset(dongle, 0x13, 1, 1)  # clear SWC UV

    def clear_SWD_UV(self, dongle):
        rgtset(dongle, 0x13, 1, 0)  # clear SWD UV

    def clear_VIN_OV(self, dongle):
        rgtset(dongle, 0x10, 1, 0)  # clear VIN OV

    def clear_VIN_UV(self, dongle):
        rgtset(dongle, 0x14, 1, 3)  # clear VIN UV

    def clear_VM_OV(self, dongle):
        rgtset(dongle, 0x13, 1, 3)  # clear SWA UV
        rgtset(dongle, 0x10, 1, 1)  # clear VIN_MGMT OV

    def clear_All_OV(self, dongle):
        rgtset(dongle, 0x10, 1, 1)  # clear VIN_MGMT OV
        rgtset(dongle, 0x10, 1, 0)  # clear VIN OV
        rgtset(dongle, 0x12, 1, 7)  # clear SWA OV
        rgtset(dongle, 0x12, 1, 5)  # clear SWC OV
        rgtset(dongle, 0x12, 1, 4)  # clear SWD OV

    def clear_All_UV(self, dongle):
        rgtset(dongle, 0x14, 1, 3)  # clear VIN UV
        rgtset(dongle, 0x13, 1, 3)  # clear SWA UV
        rgtset(dongle, 0x13, 1, 1)  # clear SWC UV
        rgtset(dongle, 0x13, 1, 0)  # clear SWD UV

    def clear_VM_VIN_SOM(self, dongle):
        rgtset(dongle, 0x11, 1, 4)
        rgtset(dongle, 0x14, 1, 4)

    def clear_CTS(self, dongle):
        rgtset(dongle, 0x11, 1, 6)  # clear Bias PG
        rgtset(dongle, 0x14, 1, 3)  # clear VIN UV

    def clear_HTW(self, dongle):
        rgtset(dongle, 0x11, 1, 7)

    def clear_LDO1p8PG(self, dongle):
        rgtset(dongle, 0x11, 1, 5)  # clear LDO1p8

    def clear_HCW(self, dongle):
        rgtset(dongle, 0x11, 1, 3)
        rgtset(dongle, 0x11, 1, 2)
        rgtset(dongle, 0x11, 1, 1)
        rgtset(dongle, 0x11, 1, 0)

    def clear_CAMP(self, dongle):
        pass

    def clear_CurLim(self, dongle):
        rgtset(dongle, 0x13, 1, 7)
        rgtset(dongle, 0x13, 1, 6)
        rgtset(dongle, 0x13, 1, 5)
        rgtset(dongle, 0x13, 1, 4)


def maskcheck(dongle):
    rgt = []
    maskindex = []
    for i in mask_rgt:
        ret, data = dongle.dsh.i2c_read_reg(i)
        rgt.append(hex(data))
        index = mask_rgt.index(i)
        str1 = str(bin(data))[2:10].zfill(8)
        for index1, s in enumerate(str1):
            if s == '1':
                maskindex.append(maskmap[index1 + index * 8])
    print(maskindex)


def statuscheck(dongle):
    rgt = []
    statusindex = ''
    for i in status_rgt:
        ret, data = dongle.dsh.i2c_read_reg(i)
        if ret != 0:
            rgt.append(data)
        elif ret == 0:
            rgt.append(hex(data))
            index = status_rgt.index(i)
            str1 = str(bin(data))[2:10].zfill(8)
            for index1, s in enumerate(str1):
                if s == '1':
                    statusindex = statusindex + ';' + str(statusmap[index1 + index * 8])

        VOUTA = el1.volmeas(1)
        VOUTB = el1.volmeas(2)
        VOUTC = el1.volmeas(3)
        VOUTD = el1.volmeas(4)
        Iin = p1.curmeas(2)
        PG = m1.meas()
        GSI = m2.meas()

        if VOUTA < 0.3 or VOUTB < 0.3 or VOUTC < 0.3 or VOUTD < 0.3 or Iin < 0.001:
            VR = 'disabled'
        else:
            VR = 'enabled'

        if PG > 1.6:
            PG = 'H'
        elif PG < 0.3:
            PG = 'L'
        else:
            PG = 'Please Check'

        if GSI > 1.6:
            GSI = 'H'
        elif GSI < 0.3:
            GSI = 'L'
        else:
            GSI = 'Please Check'

    return rgt, VR, PG, GSI, statusindex


def rgtset(dongle, rgt, set, *n):
    ret, data = dongle.dsh.i2c_read_reg(rgt)
    if set == 1:
        data = bm.set_bit(data, *n)
    elif set == 0:
        data = bm.clear_bit(data, *n)
    dongle.dsh.i2c_write_reg(rgt, data)


wb = Workbook()
ws1 = wb.active

ws1.cell(row=1, column=1, value='Item')
ws1.cell(row=1, column=2, value='Status Register')
ws1.cell(row=1, column=3, value='VR')
ws1.cell(row=1, column=4, value='PG')
ws1.cell(row=1, column=5, value='GSI')
ws1.cell(row=1, column=6, value='Status Value')
ws1.cell(row=1, column=7, value='Clear Flag')
ws1.cell(row=1, column=8, value='VR_Cleared')
ws1.cell(row=1, column=9, value='PG_Cleared')
ws1.cell(row=1, column=10, value='GSI_Cleared')

a = Piht()
e = Ej()
c = clear_event()
mask = 1

for i in range(20):
    p1.off(1)
    p1.off(2)
    time.sleep(1)
    p1.on(3.3, 1, 1)
    p1.on(12, 2, 2)
    time.sleep(0.5)

    a.unlock()
    a.GSI(True)
    a.MaskCtrl(0b10)

    a.enable(True)
    time.sleep(0.5)

    if i == 0:
        e.SWA_OV(a, mask)

    elif i == 1:
        e.SWB_OV(a, mask)

    elif i == 2:
        e.SWC_OV(a, mask)

    elif i == 3:
        e.SWD_OV(a, mask)

    elif i == 4:
        e.SWAUV(a, mask)

    elif i == 5:
        e.SWBUV(a, mask)

    elif i == 6:
        e.SWCUV(a, mask)

    elif i == 7:
        e.SWDUV(a, mask)

    elif i == 8:
        e.VINOV(a, mask)

    elif i == 9:
        e.VINUV(a, mask)

    elif i == 10:
        e.VMOV(a, mask)

    elif i == 11:
        e.All_OV(a, mask)

    elif i == 12:
        e.All_UV(a, mask)

    elif i == 13:
        e.VM_VIN_SOM(a, mask)

    elif i == 14:
        e.CTS(a, mask)

    elif i == 15:
        e.HTW(a, mask)

    elif i == 16:
        e.LDO1p8PG(a, mask)

    elif i == 17:
        e.HCW(a, mask)

    elif i == 18:
        e.CAMP(a, mask)

    elif i == 19:
        e.CurLim(a, mask)

    status, VR, PG, GSI, statusvalue = statuscheck(a)
    print('{} status: {}; VR: {}; PG: {}; GSI: {};statusvalue: {}'.format(name, status, VR, PG, GSI, statusvalue))

    a.dsh.i2c_write_reg(0x35, 0x00)

    if i == 0:
        c.clear_SWA_OV(a)

    elif i == 1:
        c.clear_SWB_OV(a)

    elif i == 2:
        c.clear_SWC_OV(a)

    elif i == 3:
        c.clear_SWD_OV(a)

    elif i == 4:
        c.clear_SWA_UV(a)

    elif i == 5:
        c.clear_SWB_UV(a)

    elif i == 6:
        c.clear_SWC_UV(a)

    elif i == 7:
        c.clear_SWD_UV(a)

    elif i == 8:
        c.clear_VIN_OV(a)

    elif i == 9:
        c.clear_VIN_UV(a)

    elif i == 10:
        c.clear_VM_OV(a)

    elif i == 11:
        c.clear_All_OV(a)

    elif i == 12:
        c.clear_All_UV(a)

    elif i == 13:
        c.clear_VM_VIN_SOM(a)

    elif i == 14:
        c.clear_CTS(a)

    elif i == 15:
        c.clear_HTW(a)

    elif i == 16:
        c.clear_LDO1p8PG(a)

    elif i == 17:
        c.clear_HCW(a)

    elif i == 18:
        c.clear_CAMP(a)

    elif i == 19:
        c.clear_CurLim(a)

    statusC, VRC, PGC, GSIC, statusvalueC = statuscheck(a)
    print('{} statusC: {}; VRC: {}; PGC: {}; GSIC: {};statusvalueC: {}'.format(name, statusC, VRC, PGC, GSIC,
                                                                               statusvalueC))
    if statusvalueC == '':
        flag = 0
    else:
        flag = statusvalueC

    ws1.cell(row=2 + i, column=1, value=name)
    ws1.cell(row=2 + i, column=2, value=(str(status[0]) + ',' + str(status[1])) + ',' + str(status[2]) + ',' + str(
        status[3]) + ',' + str(status[4]))
    ws1.cell(row=2 + i, column=3, value=VR)
    ws1.cell(row=2 + i, column=4, value=PG)
    ws1.cell(row=2 + i, column=5, value=GSI)
    ws1.cell(row=2 + i, column=6, value=statusvalue)
    ws1.cell(row=2 + i, column=7, value=flag)
    ws1.cell(row=2 + i, column=8, value=VRC)
    ws1.cell(row=2 + i, column=9, value=PGC)
    ws1.cell(row=2 + i, column=10, value=GSIC)

    wb.save(path)
