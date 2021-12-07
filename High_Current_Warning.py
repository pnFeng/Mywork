import time
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from Mylib.bench_resource import brsc
from ICs import PineHurst
from bit_manipulate import bm


class Piht:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

    def enable(self, bEnable):
        if bEnable:
            self.dsh.i2c_write_reg(0x32, 0x80)
        else:
            self.dsh.i2c_write_reg(0x32, 0x00)

    def Set_HiCurWan(self, rail, val):
        if rail == 'a':
            self.dsh.i2c_write_reg(0x1C, val << 2)
        elif rail == 'b':
            self.dsh.i2c_write_reg(0x1E, val << 2)
        elif rail == 'c':
            self.dsh.i2c_write_reg(0x1F, val << 2)

    def get_HiCurWan(self, rail):
        if rail == 'a':
            rgt = 0x1C
        elif rail == 'b':
            rgt = 0x1E
        elif rail == 'c':
            rgt = 0x1F
        ret, data = self.dsh.i2c_read_reg(rgt)
        return data

    def Get_HiCurWanflag(self, rail):
        ret, R09 = self.dsh.i2c_read_reg(0x09)
        flag = bm.get_bit(R09, 's', [3, 1, 0])
        return R09, flag

    def CurMeas(self, power, acc, rail):
        sum = 0
        if rail == 'a':
            rgt = 0x0C
        elif rail == 'b':
            rgt = 0x0E
        elif rail == 'c':
            rgt = 0x0F
        if acc == 0:
            ret, data = self.dsh.i2c_read_reg(0x6C)
            data = bm.clear_bit(data, 4)
            self.dsh.i2c_write_reg(0x6C, data)
            if power == 0:
                ret, data = self.dsh.i2c_read_reg(0x1B)
                data = bm.clear_bit(data, 6)
                self.dsh.i2c_write_reg(0x1B, data)
            elif power == 1:
                ret, data = self.dsh.i2c_read_reg(0x1B)
                data = bm.set_bit(data, 6)
                self.dsh.i2c_write_reg(0x1B, data)
        elif acc == 1:
            ret, data = self.dsh.i2c_read_reg(0x6C)
            data = bm.set_bit(data, 4)
            self.dsh.i2c_write_reg(0x6C, data)
            if power == 0:
                ret, data = self.dsh.i2c_read_reg(0x1B)
                data = bm.clear_bit(data, 6)
                self.dsh.i2c_write_reg(0x1B, data)
            elif power == 1:
                ret, data = self.dsh.i2c_read_reg(0x1B)
                data = bm.set_bit(data, 6)
                self.dsh.i2c_write_reg(0x1B, data)
        for i in range(10):
            ret, data = self.dsh.i2c_read_reg(rgt)
            ret, data1 = self.dsh.i2c_read_reg(0x2E)
            if acc == 1:
                if rail == 'a':
                    data = (data << 1) + bm.get_bit(data1, 'l', 7)[0]
                elif rail == 'b':
                    data = (data << 1) + bm.get_bit(data1, 'l', 5)[0]
                elif rail == 'c':
                    data = (data << 1) + bm.get_bit(data1, 'l', 4)[0]

            sum = data + sum
            time.sleep(0.1)
            avag = round(sum / 10, 2)
        return avag


Rail = 'c'
step = 0.125
path = '../HighCurWarning_' + Rail + '.xlsx'

p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
s1 = brsc.Scope('USB0::0x05FF::0x1023::3808N60392::INSTR')
m1 = brsc.DMM('USB0::0x2A8D::0x1401::MY57215694::0::INSTR')
m2 = brsc.DMM('USB0::0x2A8D::0x1401::MY57218437::0::INSTR')
p1.on(5, 3, 1)

wb = Workbook()
ws1 = wb.active
ws1.cell(row=1, column=1, value='Setting')
ws1.cell(row=1, column=2, value='load')
ws1.cell(row=1, column=3, value='ADC Meas')
ws1.cell(row=1, column=4, value='ADC Meas Value(A)')
ws1.cell(row=1, column=5, value='R09 Status')
ws1.cell(row=1, column=6, value='Target(A)')

a = Piht()
a.enable(True)
time.sleep(1)

for i in range(1, 13, 1):
    a.Set_HiCurWan(Rail, i)
    el = 0
    while 1:
        el1.cc(el, 2)
        el1.on([2])
        R09, f = a.Get_HiCurWanflag(Rail)
        print(i, el, f)
        if f != '000':
            time.sleep(0.5)
            Reading = a.CurMeas(0, 0, Rail)
            el1.off([2])
            time.sleep(0.1)
            a.dsh.i2c_write_reg(0x14, 0x01)  # Global Clear
            ws1.cell(row=i + 1, column=1, value='0x' + str(hex(a.get_HiCurWan(Rail)))[2:].zfill(2))
            ws1.cell(row=i + 1, column=2, value=el)
            ws1.cell(row=i + 1, column=3, value=Reading)
            ws1.cell(row=i + 1, column=4, value=Reading * step)
            ws1.cell(row=i + 1, column=5, value='0x' + str(hex(R09))[2:].zfill(2))
            ws1.cell(row=i + 1, column=6, value=i * step)
            break
        elif el >= 8:
            break
        else:
            el += 0.05

wb.save(path)

el1.off([2])
