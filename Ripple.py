import time
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from Mylib.bench_resource import brsc
from ICs import PineHurst
from bit_manipulate import bm


class Piht:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

        # Un-lock DIMM vendor region(0x40~0x6f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x70~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        ret, data = self.dsh.i2c_read_reg(0x2F)
        assert ret == 0
        data = bm.set_bit(data, 2)
        self.dsh.i2c_write_reg(0x2F, data)

    def enable(self, bEnable):
        ret, data = self.dsh.i2c_read_reg(0x32)
        if bEnable:
            data = bm.set_bit(data, 7)
            self.dsh.i2c_write_reg(0x32, data)
        else:
            data = bm.clear_bit(data, 7)
            self.dsh.i2c_write_reg(0x32, data)


Rail = 'A'
path = r'D:\DDR_PMIC\Pinehurst\SK_ReducedCap\UDIMM\SW' + Rail
xlsx = path + '\Ripple.xlsx'
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
s1 = brsc.Scope('USB0::0x05FF::0x1023::5004N60546::0::INSTR')

# p1.off(1)
time.sleep(1)
# p1.on(5, 2, 1)

p1.on(3.3, 2, 1)
p1.on(12, 2, 2)

wb = Workbook()
ws1 = wb.active

ws1.cell(row=1, column=1, value='LOAD')
ws1.cell(row=1, column=2, value='RIPPLE')
ws1.cell(row=1, column=3, value='FSW')

a = Piht()
# bm.bit_manipulate(0x82, a, (6, 0), (5, 1), (4, 1))  # enable OOA
# bm.bit_manipulate(0xac, a, (2, 1), (1, 0), (0, 1))  # Ramp
# bm.bit_manipulate(0x2F, a, (4, 1))
# bm.bit_manipulate(0x2A, a, (0, 0))  # SWC 750kHz
bm.bit_manipulate(0x2F, a, (4, 1))

if Rail == 'A':
    # a.dsh.i2c_write_reg(0x2f, 0x42)  # RAILA
    # a.dsh.util_i2c_modify_reg_bits(0x2b, [(5, 1)])  # Range
    # a.dsh.i2c_write_reg(0x21, 0xFE)  # set VID
    # print(a.dsh.i2c_read_reg(0x27))
    # a.dsh.i2c_write_reg(0x2F, 0x46)
    # print(a.dsh.i2c_read_reg(0x2F))
    channel = 3
    max = 5
    load = [0, 0.1, 0.2, 0.5, 1, 2, 3, 4]
elif Rail == 'B':
    # a.dsh.i2c_write_reg(0x2f, 0x12)  # RAILB
    # bm.bit_manipulate(0x2A, a, (5, 0), (4, 1))  # FSW
    # bm.bit_manipulate(0x83, a, (1, 1))  # Filter
    # bm.bit_manipulate(0xb2, a, (7, 1), (6, 1))
    channel = 2
    max = 5
    load = [0, 0.01, 0.1, 0.5]
elif Rail == 'C':
    # a.dsh.i2c_write_reg(0x2A, 0x89)
    # a.dsh.i2c_write_reg(0x2f, 0x0A)  # RAILC
    # a.dsh.util_i2c_modify_reg_bits(0x2b, [(0, 1)])  # Range
    # a.dsh.i2c_write_reg(0x27, 0xFE)  # set VID
    # print(a.dsh.i2c_read_reg(0x27))
    # a.dsh.i2c_write_reg(0x2F, 0x0E)
    # print(a.dsh.i2c_read_reg(0x2F))
    channel = 3
    max = 1
    load = [0, 0.1, 0.2, 0.5, 1]

# a.enable(True)

j = 0
for i in load:
    j += 1
    if i < 0.1:
        s1.scope.write('vbs app.Acquisition.Horizontal.HorScale = 0.1e-3')
    elif 0.1 <= i < 0.3:
        s1.scope.write('vbs app.Acquisition.Horizontal.HorScale = 5e-6')
    elif 0.3 <= i < 0.5:
        s1.scope.write('vbs app.Acquisition.Horizontal.HorScale = 5e-6')
    else:
        s1.scope.write('vbs app.Acquisition.Horizontal.HorScale = 2e-6')
    el1.cc(i, channel)
    print(i)
    el1.on([channel])
    time.sleep(0.5)
    s1.scope.write('vbs app.acquisition.triggermode = "Auto"')
    s1.scope.write('vbs app.measure.clearsweeps')
    time.sleep(3)

    # c1 = s1.scope.query('vbs? return = app.measure.p3.max.result.value')
    c1 = float(s1.scope.query('vbs? return = app.measure.p3.max.result.value')) * 1000
    # c1 = float(s1.scope.query('vbs? return = app.measure.p3.max.result.value')[3:]) * 1000
    # FSW = float(s1.scope.query('vbs? return = app.measure.p4.mean.result.value')[3:]) / 1000
    # min = s1.scope.query('vbs? return = app.measure.p2.min.result.value')

    s1.SCap(path + '\load_' + str(i) + '.png')

    ws1.cell(row=j + 1, column=1, value=i)
    ws1.cell(row=j + 1, column=2, value=round(c1, 3))
    # ws1.cell(row=j + 1, column=3, value=FSW)

wb.save(xlsx)

el1.off([channel])
