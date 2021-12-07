import time
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from Mylib.bench_resource import brsc
from ICs import PineHurst
from Mylib.bit_manipulate import bm
from openpyxl.styles import colors, Font
import atexit


class Piht:
    def __init__(self):
        self.dsh = PineHurst()
        ret = self.dsh.connect()
        assert ret == True

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

        # Un-lock DIMM vendor region(0x40~0x6f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x70~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        bm.bit_manipulate(self, 0x2F, (2, 1))

    def enable(self, bEnable):
        if bEnable:
            bm.bit_manipulate(self, 0x32, (7, 1))
        else:
            bm.bit_manipulate(self, 0x32, (7, 0))


path = './../' + 'EFF.xlsx'
# global p1, el1, m1, m2
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
m1 = brsc.DMM('USB0::0x2A8D::0x1401::MY57215694::0::INSTR')
m2 = brsc.DMM('USB0::0x2A8D::0x1401::MY57218437::0::INSTR')


@atexit.register
def clean():
    global p1, el1, m1, m2
    del p1, el1, m1, m2
    print('=' * 40 + 'CLEAN' + '=' * 40)


p1.on(3, 2, 1)
wb = Workbook()
ws1 = wb.active

a = Piht()
bm.bit_manipulate(a, 0x82, (4, 1))  # enable OOA
bm.bit_manipulate(a, 0x83, (1, 0), (0, 1))  # OOA

Rail = 'A'

if Rail == 'A':
    bm.bit_manipulate(a, 0x29, (5, 0), (4, 1))
    a.dsh.i2c_write_reg(0x2f, 0x42)  # RAILA
    channel = 1
    max = 5
    load = [0.01 * x for x in range(10)] + \
           [0.1 * x for x in range(10)] + \
           [1 + 0.25 * x for x in range(10)]
elif Rail == 'B':
    # bm.bit_manipulate(a, 0x2A,(5, 1), (4, 0))  # FSW
    # bm.bit_manipulate(a, 0x83, (1, 1))  # Filter
    # bm.bit_manipulate(a, 0xb2, (7, 1), (6, 1))
    # bm.bit_manipulate(a, 0x6D, (1, 1))
    # a.dsh.i2c_write_reg(0x25, 0x28)
    a.dsh.i2c_write_reg(0x2f, 0x12)  # RAILB
    channel = 2
    max = 5
    load = [
        0, 0.01, 0.02, 0.03, 0.04, 0.05, 0.06, 0.07, 0.08, 0.09, 0.1, 0.15,
        0.2, 0.25, 0.3, 0.35, 0.4, 0.45, 0.5, 0.6, 0.7, 0.8, 0.9, 1, 1.25, 1.5,
        1.75, 2, 2.25, 2.5, 2.75, 3, 3.25, 3.5, 3.75, 4, 4.25, 4.5, 4.75, 5
    ]
elif Rail == 'C':
    # a.dsh.i2c_write_reg(0x2A, 0x8a)
    a.dsh.i2c_write_reg(0x2f, 0x0E)  # RAILC
    channel = 3
    max = 1
    load = [
        0, 0.01, 0.02, 0.03, 0.04, 0.05, 0.06, 0.07, 0.08, 0.09, 0.1, 0.15,
        0.2, 0.25, 0.3, 0.35, 0.4, 0.45, 0.5, 0.6, 0.7, 0.8, 0.9, 1
    ]

print(a.dsh.i2c_read_reg(0x2A))

step1 = 0.1
step2 = 0.25
group = 0
maxcell = 0
i, j = -step1, 0

j = 0
for i in load:
    j += 1
    el1.cc(i, channel)
    print(i)
    el1.on([channel])
    time.sleep(0.5)

    iout = el1.curmeas(channel)
    v1 = m1.meas()
    v2 = m2.meas()
    Iin_p = p1.curmeas(1)

    power_in = v1 * Iin_p
    power_out = v2 * iout
    Eff = 100 * power_out / power_in
    if maxcell < Eff:
        maxcell = Eff
        red_c = 5 + 8 * group
        red_r = j + 1

    ws1.cell(row=j + 1, column=1 + 8 * group, value=v1)
    ws1.cell(row=j + 1, column=2 + 8 * group, value=Iin_p)
    ws1.cell(row=j + 1, column=3 + 8 * group, value=v2)
    ws1.cell(row=j + 1, column=4 + 8 * group, value=iout)
    ws1.cell(row=j + 1, column=5 + 8 * group, value=Eff)

    ws1.cell(row=1, column=1 + 8 * group, value='Vin_Bulk')
    ws1.cell(row=1, column=2 + 8 * group, value='Iin_Bulk')
    ws1.cell(row=1, column=3 + 8 * group, value='VOUT')
    ws1.cell(row=1, column=4 + 8 * group, value='IOUT')
    ws1.cell(row=1, column=5 + 8 * group, value='Efficiency')

ft = Font(color='00003366', size=12)
c = ws1.cell(row=red_r, column=red_c)
c.font = ft

wb.save(path)

el1.off([channel])
