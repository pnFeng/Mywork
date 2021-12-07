# -*- coding: utf-8 -*-

import visa
import sys
import time
from openpyxl import Workbook
import dolphin
from ICs import DolphinSH
from bit_manipulate import bm
from Mylib.bench_resource import brsc


class PH:
    def __init__(self):
        self.dsh = DolphinSH()
        ret = self.dsh.connect()
        if ret is False:
            print("Failed to connect")
            sys.exit(-1)

        self.dsh.helper_mark_global_vars(lid=self.dsh.LID, hid=7)
        self.dsh.set_voltage(1.0)
        self.dsh.set_i2c_bitrate(100)

        # Un-lock DIMM vendor region(0x40~0x7f)
        self.dsh.i2c_write_reg(0x37, 0x73)
        self.dsh.i2c_write_reg(0x38, 0x94)
        self.dsh.i2c_write_reg(0x39, 0x40)

        # Un-lock PMIC vendor region(0x80~0xff)
        self.dsh.i2c_write_reg(0x37, 0x79)
        self.dsh.i2c_write_reg(0x38, 0xbe)
        self.dsh.i2c_write_reg(0x39, 0x10)

        # Diable OVP
        self.dsh.i2c_write_reg(0x56, 0xF0)

    def enable(self):
        # enable
        ret, data = self.dsh.i2c_read_reg(0x2F)
        print('data', data, 'ret', ret)
        data = bm.set_bit(data, 3)
        self.dsh.i2c_write_reg(0x2F, data)
        self.dsh.i2c_write_reg(0x32, 0x80)

    def RangeSelect(self, ch, rg):
        if ch == 'a':
            if rg == 0:
                ret, data = self.dsh.i2c_read_reg(0x6B)
                data = data & 0x3F
                self.dsh.i2c_write_reg(0x6B, data)
                ret, data = self.dsh.i2c_read_reg(0x2B)
                data = data & 0xDF
                self.dsh.i2c_write_reg(0x2B, data)
                ret, data = self.dsh.i2c_read_reg(0x6D)
                data = data & 0xF7
                self.dsh.i2c_write_reg(0x6D, data)
            elif rg == 1:
                ret, data = self.dsh.i2c_read_reg(0x6B)
                data = data & 0x3F
                self.dsh.i2c_write_reg(0x6B, data)
                ret, data = self.dsh.i2c_read_reg(0x2B)
                data = data & 0xDF
                self.dsh.i2c_write_reg(0x2B, data)
                ret, data = self.dsh.i2c_read_reg(0x6D)
                data = data | 0x08
                self.dsh.i2c_write_reg(0x6D, data)
            elif rg == 2:
                ret, data = self.dsh.i2c_read_reg(0x6B)
                data = data & 0x3F
                self.dsh.i2c_write_reg(0x6B, data)
                ret, data = self.dsh.i2c_read_reg(0x2B)
                data = data | 0x20
                self.dsh.i2c_write_reg(0x2B, data)
                ret, data = self.dsh.i2c_read_reg(0x6D)
                data = data & 0xF7
                self.dsh.i2c_write_reg(0x6D, data)
            elif rg == 3:
                ret, data = self.dsh.i2c_read_reg(0x6B)
                data = data & 0x3F
                self.dsh.i2c_write_reg(0x6B, data)
                ret, data = self.dsh.i2c_read_reg(0x2B)
                data = data | 0x20
                self.dsh.i2c_write_reg(0x2B, data)
                ret, data = self.dsh.i2c_read_reg(0x6D)
                data = data | 0x08
                self.dsh.i2c_write_reg(0x6D, data)
            elif rg == 4:
                ret, data = self.dsh.i2c_read_reg(0x6B)
                data = data | 0x80 & 0xBF
                self.dsh.i2c_write_reg(0x6B, data)
                ret, data = self.dsh.i2c_read_reg(0x2B)
                data = data & 0xDF
                self.dsh.i2c_write_reg(0x2B, data)
                ret, data = self.dsh.i2c_read_reg(0x6D)
                data = data & 0xF7
                self.dsh.i2c_write_reg(0x6D, data)
            elif rg == 5:
                ret, data = self.dsh.i2c_read_reg(0x6B)
                data = data | 0x80 & 0xBF
                self.dsh.i2c_write_reg(0x6B, data)
                ret, data = self.dsh.i2c_read_reg(0x2B)
                data = data | 0x20
                self.dsh.i2c_write_reg(0x2B, data)
                ret, data = self.dsh.i2c_read_reg(0x6D)
                data = data & 0xF7
                self.dsh.i2c_write_reg(0x6D, data)
            elif rg == 6:
                ret, data = self.dsh.i2c_read_reg(0x6B)
                data = data | 0x80 & 0xBF
                self.dsh.i2c_write_reg(0x6B, data)
                ret, data = self.dsh.i2c_read_reg(0x2B)
                data = data | 0x20
                self.dsh.i2c_write_reg(0x2B, data)
                ret, data = self.dsh.i2c_read_reg(0x6D)
                data = data | 0x08
                self.dsh.i2c_write_reg(0x6D, data)
            elif rg == 7:
                ret, data = self.dsh.i2c_read_reg(0x6B)
                data = data | 0xC0
                self.dsh.i2c_write_reg(0x6B, data)
                ret, data = self.dsh.i2c_read_reg(0x2B)
                data = data & 0xDF
                self.dsh.i2c_write_reg(0x2B, data)
                ret, data = self.dsh.i2c_read_reg(0x6D)
                data = data & 0xF7
                self.dsh.i2c_write_reg(0x6D, data)
            elif rg == 8:
                ret, data = self.dsh.i2c_read_reg(0x6B)
                data = data | 0xC0
                self.dsh.i2c_write_reg(0x6B, data)
                ret, data = self.dsh.i2c_read_reg(0x2B)
                data = data | 0x20
                self.dsh.i2c_write_reg(0x2B, data)
                ret, data = self.dsh.i2c_read_reg(0x6D)
                data = data & 0xF7
                self.dsh.i2c_write_reg(0x6D, data)
            elif rg == 9:
                ret, data = self.dsh.i2c_read_reg(0x6B)
                data = data | 0xC0
                self.dsh.i2c_write_reg(0x6B, data)
                ret, data = self.dsh.i2c_read_reg(0x2B)
                data = data | 0x20
                self.dsh.i2c_write_reg(0x2B, data)
                ret, data = self.dsh.i2c_read_reg(0x6D)
                data = data | 0x08
                self.dsh.i2c_write_reg(0x6D, data)

    def VIDSetting(self, ch, VID, cre):
        if ch == 'a':
            self.value = VID << 1 | cre
            self.dsh.i2c_write_reg(0x21, self.value)

    def ADCConfig(self, ch):
        if ch == 'a':
            self.dsh.i2c_write_reg(0x30, 0x80)
        elif ch == 'b':
            self.dsh.i2c_write_reg(0x30, 0x90)
        elif ch == 'c':
            self.dsh.i2c_write_reg(0x30, 0x98)

    def ADCrslt(self):
        ret, data1 = self.dsh.i2c_read_reg(0x31)
        ret, data2 = self.dsh.i2c_read_reg(0x72)
        return data1, data2

    def enableFLT(self, enable):
        if enable == 1:
            ret, data = self.dsh.i2c_read_reg(0xc7)
            data = bm.set_bit(data, 6)
            self.dsh.i2c_write_reg(0xc7, data)
            ret, data = self.dsh.i2c_read_reg(0xc7)
            print('0xc7_1', hex(data))
        elif enable == 0:
            ret, data = self.dsh.i2c_read_reg(0xc7)
            data = bm.clear_bit(data, 6)
            self.dsh.i2c_write_reg(0xc7, data)
            ret, data = self.dsh.i2c_read_reg(0xc7)
            print
            '0xc7_0', hex(data)


path = './../' + 'ADC_A.xlsx'
p1 = brsc.DPS('USB0::0x2A8D::0x1202::MY58270817::0::INSTR')
el1 = brsc.Eload('USB0::0x0A69::0x083E::636002002010::0::INSTR')
m1 = brsc.DMM('USB0::0x2A8D::0x1401::MY57215694::0::INSTR')
m2 = brsc.DMM('USB0::0x2A8D::0x1401::MY57218437::0::INSTR')
p1.on(5, 2, 1)

wb = Workbook()
ws1 = wb.active

a = PH()
a.enable()
a.ADCConfig('a')
a.enableFLT(0)
time.sleep(1)

for j in range(1):  # 10
    a.RangeSelect('a', j)
    time.sleep(0.5)
    ret, data1 = a.dsh.i2c_read_reg(0x6B)
    ret, data2 = a.dsh.i2c_read_reg(0x2B)
    ret, data3 = a.dsh.i2c_read_reg(0x6D)
    print
    'RangeSetting', hex(data1), hex(data2), hex(data3)

    data1 = bin(data1)[2:].zfill(8)
    data2 = bin(data2)[2:].zfill(8)
    data3 = bin(data3)[2:].zfill(8)

    ws1.cell(1, 1 + 12 * j, j)
    ws1.cell(1, 2 + 12 * j, '0x6B[7:6] = ' + data1[0:2])
    ws1.cell(1, 3 + 12 * j, '0x2B[5] = ' + data2[2])
    ws1.cell(1, 4 + 12 * j, '0x6D[3] = ' + data3[4])

    localtime = time.asctime(time.localtime(time.time()))
    ws1.cell(2, 1 + 12 * j, '#')
    ws1.cell(2, 2 + 12 * j, 'Multimeter VOUT(V)')
    ws1.cell(2, 3 + 12 * j, 'Setting 0x21[7:1]')
    ws1.cell(2, 4 + 12 * j, 'ADC_8bit_Average')
    ws1.cell(2, 5 + 12 * j, 'ADC_10bit_Average')
    ws1.cell(2, 6 + 12 * j, 'ADC_8bit Value(mV)')
    ws1.cell(2, 7 + 12 * j, 'ADC_10bit Value(mV)')
    ws1.cell(2, 8 + 12 * j, 'ADC_8bit Delta_LSB')
    ws1.cell(2, 9 + 12 * j, 'ADC_10bit Delta_LSB')
    for x in range(1, 11, 1):
        ws1.cell(2, 9 + x + 12 * j, str(x) + '_times_10bit')

    n = 128
    for i in range(n):
        print
        i, j
        a.VIDSetting('a', i, 0)
        time.sleep(0.5)
        ret, data = a.dsh.i2c_read_reg(0x21)
        meas1 = m2.meas()

        Sum, Sum1 = 0, 0
        for k in range(10):
            ADC1, ADC2 = a.ADCrslt()
            ADCData = (ADC1 * 15)
            Sum += ADC1
            ADC_10bit = (ADC1 << 2) + int(bm.get_bit(ADC2, 's', [1, 0]), 2)
            ws1.cell(i + 3, 10 + k + 12 * j, hex(ADC_10bit))
            Sum1 += ADC_10bit
            time.sleep(0.05)
        ADCavg = int(round(Sum / 10))
        ADCavg_10bit = int(round(Sum1 / 10))

        targetADC_8bit = int(round(meas1 / 0.015))
        LSB_8bit = targetADC_8bit - ADCavg

        targetADC_10bit = int(round(meas1 / 0.00375))
        LSB_10bit = targetADC_10bit - ADCavg_10bit

        if j == 0:
            target = 800 + 5 * i
        elif j == 1:
            target = 600 + 3.75 * i
        elif j == 2:
            target = 600 + 5 * i
        elif j == 3:
            target = 600 + 5 * i
        elif j == 4:
            target = 1500 + 5 * i
        elif j == 5:
            target = 1300 + 5 * i
        elif j == 6:
            target = 1300 + 5 * i
        elif j == 7:
            target = 2200 + 5 * i
        elif j == 8:
            target = 2000 + 5 * i
        elif j == 9:
            target = 2000 + 5 * i

        error = 100 * (1000 * meas1 - target) / target

        ADCData = (ADCavg * 15)
        ADCData_10bit = (ADCavg_10bit * 3.75)

        print
        ADCData, ADCData_10bit

        ws1.cell(i + 3, 1 + 12 * j, i)
        ws1.cell(i + 3, 2 + 12 * j, meas1)
        ws1.cell(i + 3, 3 + 12 * j, hex(data))
        ws1.cell(i + 3, 4 + 12 * j, hex(ADCavg))
        ws1.cell(i + 3, 5 + 12 * j, hex(ADCavg_10bit))
        ws1.cell(i + 3, 6 + 12 * j, ADCData)
        ws1.cell(i + 3, 7 + 12 * j, ADCData_10bit)
        ws1.cell(i + 3, 8 + 12 * j, LSB_8bit)
        ws1.cell(i + 3, 9 + 12 * j, LSB_10bit)

wb.save(path)
